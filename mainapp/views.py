from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Sum
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.utils.html import strip_tags
from django.db.models import Sum, Q
from django.db import transaction
from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.conf import settings
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import Location, TypeEquipment, EquipmentLocation, Photo, Order,\
    OrderItem, CommonEquipmentLocation, Application, History, Feedback, Department,\
        UserDepartment, ApplicationApproval, DepartmentTypeEquipment
from .history_utils import HistoryService
from django.contrib.admin.views.decorators import staff_member_required
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import timedelta, datetime
from .email_utils import EmailService


def feedback_form(request, app_id):
    """Страница формы обратной связи по заявке"""
    try:
        # Проверяем, авторизован ли пользователь
        if not request.user.is_authenticated:
            return redirect('/')
        
        application = Application.objects.get(id=app_id, id_user=request.user)
        
        # Проверяем, что заявка завершена по времени
        if application.date_end and application.date_end > timezone.now():
            # Если мероприятие еще не завершено, перенаправляем на главную с сообщением
            messages.error(request, 'Мероприятие еще не завершено')
            return redirect('/')
        
        # Проверяем, не оставлял ли пользователь уже отзыв
        existing_feedback = Feedback.objects.filter(
            id_application=application,
            id_user=request.user
        ).exists()
        
        if existing_feedback:
            return render(request, 'mainapp/feedback_already.html', {
                'application': application
            })
        
        context = {
            'application_id': application.id,
            'application_name': application.name,
        }
        return render(request, 'mainapp/feedback_form.html', context)
        
    except Application.DoesNotExist:
        # Логируем ошибку
        print(f"Заявка с ID {app_id} не найдена для пользователя {request.user.username}")
        return JsonResponse({
            'success': False,
            'error': 'Заявка не найдена'
        }, status=404)
        
        
def get_all_equipment_types_for_application(application):
    """
    Получает все типы оборудования из ВСЕХ заказов заявки
    """
    equipment_types = set()
    
    orders = Order.objects.filter(id_application=application)
    for order in orders:
        items = OrderItem.objects.filter(order=order)
        for item in items:
            if item.equipment_location:
                equipment_types.add(item.equipment_location.id_types_equipments.id)
            elif item.common_equipment_location:
                equipment_types.add(item.common_equipment_location.id_types_equipments.id)
    
    return equipment_types


def get_departments_for_equipment_types(equipment_types):
    """
    Находит подразделения, которые должны согласовывать данные типы оборудования
    """
    departments = Department.objects.filter(
        is_approval_required=True
    ).filter(
        department_type_equipment__id_type_equipment__id__in=equipment_types
    ).distinct()
    
    return departments


def get_equipment_types_from_application(application):
    """
    Получает все типы оборудования, используемые в заявке
    """
    equipment_types = set()
    
    orders = Order.objects.filter(id_application=application)
    for order in orders:
        items = OrderItem.objects.filter(order=order)
        for item in items:
            if item.equipment_location:
                equipment_types.add(item.equipment_location.id_types_equipments.id)
            elif item.common_equipment_location:
                equipment_types.add(item.common_equipment_location.id_types_equipments.id)
    
    return equipment_types

def update_application_approvals(application, equipment_types):
    """
    Обновляет список согласований для заявки на основе текущих типов оборудования
    """
    if not equipment_types:
        # Если нет типов оборудования, удаляем все согласования
        deleted_count, _ = ApplicationApproval.objects.filter(id_application=application).delete()
        print(f"Удалено {deleted_count} согласований для заявки #{application.id} (нет оборудования)")
        return 0
    
    # Получаем подразделения, которые должны согласовывать
    required_departments = get_departments_for_equipment_types(equipment_types)
    required_department_ids = set(required_departments.values_list('id', flat=True))
    
    # Получаем текущие согласования
    current_approvals = ApplicationApproval.objects.filter(id_application=application)
    current_department_ids = set(current_approvals.values_list('id_department_id', flat=True))
    
    # Добавляем новые согласования
    departments_to_add = required_departments.filter(id__in=required_department_ids - current_department_ids)
    added_count = 0
    for dept in departments_to_add:
        ApplicationApproval.objects.create(
            id_department=dept,
            id_application=application,
            is_agreed=False,
            comment=f'Автоматически добавлено при обновлении заявки'
        )
        added_count += 1
        print(f"Добавлено согласование для подразделения {dept.name}")
    
    # Удаляем согласования для подразделений, которых больше нет
    departments_to_remove = current_approvals.filter(id_department_id__in=current_department_ids - required_department_ids)
    removed_count = departments_to_remove.delete()[0]
    if removed_count:
        print(f"Удалено {removed_count} согласований для подразделений, больше не участвующих в заявке")
    
    return added_count


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_feedback(request):
    """
    Создание отзыва/обратной связи и отправка email уведомления
    """
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        comment = data.get('comment', '').strip()
        
        # Сбор дополнительных данных
        extra_data = {}
        
        # Информация о браузере
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        extra_data['user_agent'] = user_agent
        
        # Определение браузера и ОС (упрощённо)
        if 'Chrome' in user_agent and 'Edg' not in user_agent:
            extra_data['browser'] = 'Chrome'
        elif 'Firefox' in user_agent:
            extra_data['browser'] = 'Firefox'
        elif 'Safari' in user_agent and 'Chrome' not in user_agent:
            extra_data['browser'] = 'Safari'
        elif 'Edg' in user_agent:
            extra_data['browser'] = 'Edge'
        elif 'Opera' in user_agent or 'OPR' in user_agent:
            extra_data['browser'] = 'Opera'
        else:
            extra_data['browser'] = 'Other'
        
        # Определение ОС
        if 'Windows' in user_agent:
            extra_data['os'] = 'Windows'
        elif 'Mac' in user_agent:
            extra_data['os'] = 'macOS'
        elif 'Linux' in user_agent:
            extra_data['os'] = 'Linux'
        elif 'Android' in user_agent:
            extra_data['os'] = 'Android'
        elif 'iPhone' in user_agent or 'iPad' in user_agent:
            extra_data['os'] = 'iOS'
        else:
            extra_data['os'] = 'Other'
        
        # Язык браузера
        extra_data['language'] = request.META.get('HTTP_ACCEPT_LANGUAGE', '')[:50]
        
        # IP адрес
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            extra_data['ip'] = x_forwarded_for.split(',')[0].strip()
        else:
            extra_data['ip'] = request.META.get('REMOTE_ADDR', '')
        
        # Реферер (откуда пришёл пользователь)
        extra_data['referer'] = request.META.get('HTTP_REFERER', '')
        
        # Текущий URL
        extra_data['url'] = request.META.get('HTTP_REFERER', '')
        
        # ID заявки, если передана
        application_id = data.get('application_id')
        if application_id:
            extra_data['application_id'] = application_id
        
        # Проверка на пустой комментарий
        if not comment:
            return JsonResponse({
                'success': False,
                'error': 'Текст сообщения не может быть пустым'
            }, status=400)
        
        # Создаем отзыв с дополнительными данными
        feedback = Feedback.objects.create(
            id_user=request.user,
            name=name if name else None,
            comment=comment,
            data=extra_data
        )
        
        # Логируем действие
        HistoryService.add_entry(
            request.user,
            f"Добавлен отзыв #{feedback.id}: {name if name else 'Без темы'}"
        )
        
        # ========== ОТПРАВКА EMAIL УВЕДОМЛЕНИЯ ==========
        try:
            from .email_utils import EmailService
            
            # Формируем тему письма
            if name:
                subject = f"📝 Обратная связь от {request.user.username}: {name[:50]}"
            else:
                subject = f"📝 Обратная связь от {request.user.username}"
            
            # Формируем HTML письмо
            from django.template.loader import render_to_string
            from django.utils.html import strip_tags
            from django.conf import settings
            
            context = {
                'feedback_id': feedback.id,
                'user': request.user,
                'user_name': request.user.get_full_name() or request.user.username,
                'user_email': request.user.email,
                'feedback_name': name if name else 'Без темы',
                'feedback_comment': comment,
                'created_at': feedback.created_at,
                'admin_url': getattr(settings, 'SITE_URL', 'http://127.0.0.1:8000')
            }
            
            html_message = render_to_string('mainapp/email/feedback_notification.html', context)
            plain_message = strip_tags(html_message)
            
            # Отправляем всем получателям из группы MailReciver
            receivers = EmailService.get_mail_receivers()
            
            if receivers:
                for receiver in receivers:
                    try:
                        send_mail(
                            subject=subject,
                            message=plain_message,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[receiver.email],
                            html_message=html_message,
                            fail_silently=False
                        )
                        print(f"Уведомление отправлено на {receiver.email}")
                    except Exception as e:
                        print(f"Ошибка отправки на {receiver.email}: {e}")
            else:
                print("Нет получателей в группе MailReciver")
                
        except Exception as e:
            print(f"Ошибка отправки email уведомления: {e}")
            # Не блокируем создание отзыва, если письмо не отправилось
        
        return JsonResponse({
            'success': True,
            'message': 'Спасибо за обратную связь! Ваше сообщение отправлено.',
            'feedback': {
                'id': feedback.id,
                'name': feedback.name,
                'comment': feedback.comment,
                'created_at': feedback.created_at.isoformat(),
                'created_at_formatted': feedback.created_at.strftime('%d.%m.%Y %H:%M:%S')
            }
        })
        
    except Exception as e:
        print(f"Ошибка создания отзыва: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def submit_event_feedback(request):
    """
    Отправка обратной связи по завершённому мероприятию
    """
    try:
        data = json.loads(request.body)
        
        application_id = data.get('application_id')
        event_rating = data.get('event_rating')
        event_feedback = data.get('event_feedback', '').strip()
        comment = data.get('comment', '').strip()
        
        # Валидация
        if not application_id:
            return JsonResponse({
                'success': False,
                'error': 'Не указана заявка'
            }, status=400)
        
        if not event_rating or event_rating < 1 or event_rating > 10:
            return JsonResponse({
                'success': False,
                'error': 'Пожалуйста, оцените мероприятие от 1 до 10'
            }, status=400)
        
        if not event_feedback:
            return JsonResponse({
                'success': False,
                'error': 'Пожалуйста, напишите отзыв о проведении мероприятия'
            }, status=400)
        
        if not comment:
            return JsonResponse({
                'success': False,
                'error': 'Пожалуйста, оставьте комментарий'
            }, status=400)
        
        # Получаем заявку
        application = Application.objects.get(id=application_id, id_user=request.user)
        
        # Проверяем, что мероприятие завершено
        if application.date_end and application.date_end > timezone.now():
            return JsonResponse({
                'success': False,
                'error': 'Мероприятие еще не завершено'
            }, status=400)
        
        # Проверяем, не оставлял ли пользователь уже отзыв
        existing_feedback = Feedback.objects.filter(
            id_application=application,
            id_user=request.user
        ).exists()
        
        if existing_feedback:
            return JsonResponse({
                'success': False,
                'error': 'Вы уже оставляли отзыв на эту заявку'
            }, status=400)
        
        # Создаём запись в обратной связи
        feedback = Feedback.objects.create(
            id_user=request.user,
            id_application=application,
            name=f"Отзыв на заявку #{application.id}",
            comment=comment,
            data={
                'event_rating': event_rating,
                'event_feedback': event_feedback,
                'submitted_at': timezone.now().isoformat()
            }
        )
        
        # Меняем статус заявки на "completed" (Завершена)
        old_status = application.status
        application.status = 'completed'
        application.save(update_fields=['status'])
        
        # Логируем действие
        HistoryService.add_entry(
            request.user,
            f"Оставлен отзыв на заявку #{application.id} (оценка: {event_rating}/10) и статус изменён на 'Завершена'"
        )
        
        # Отправляем уведомление администраторам
        try:
            from .email_utils import EmailService
            
            # Формируем контекст для письма
            email_context = {
                'feedback_id': feedback.id,
                'application_id': application.id,
                'application_name': application.name,
                'user': request.user,
                'user_name': request.user.get_full_name() or request.user.username,
                'user_email': request.user.email,
                'event_rating': event_rating,
                'event_feedback': event_feedback,
                'comment': comment,
                'created_at': feedback.created_at,
                'old_status': old_status,
                'new_status': 'completed',
                'admin_url': getattr(settings, 'SITE_URL', 'http://127.0.0.1:8000')
            }
            
            # Рендерим HTML письма
            html_message = render_to_string('mainapp/email/event_feedback_notification.html', email_context)
            plain_message = strip_tags(html_message)
            
            subject = f"📝 Отзыв на заявку #{application.id} от {request.user.username}"
            
            receivers = EmailService.get_mail_receivers()
            if receivers:
                for receiver in receivers:
                    send_mail(
                        subject=subject,
                        message=plain_message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[receiver.email],
                        html_message=html_message,
                        fail_silently=False
                    )
        except Exception as e:
            print(f"Ошибка отправки email уведомления: {e}")
        
        return JsonResponse({
            'success': True,
            'message': 'Спасибо за отзыв! Статус заявки изменён на "Завершена".',
            'feedback': {
                'id': feedback.id,
                'rating': event_rating,
                'created_at': feedback.created_at.isoformat()
            }
        })
        
    except Application.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Заявка не найдена'
        }, status=404)
    except Exception as e:
        print(f"Ошибка: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
                
        
@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def register_view(request):
    """Регистрация нового пользователя (по email)"""
    try:
        data = json.loads(request.body)
        email = data.get('email', '').strip().lower()
        password = data.get('password')
        password_confirm = data.get('password_confirm')
        first_name = data.get('first_name', '')
        last_name = data.get('last_name', '')
        
        # Валидация email
        if not email:
            return JsonResponse({
                'success': False,
                'error': 'Email обязателен для заполнения'
            }, status=400)
        
        try:
            validate_email(email)
        except ValidationError:
            return JsonResponse({
                'success': False,
                'error': 'Введите корректный email адрес'
            }, status=400)
        
        # Проверка пароля
        if not password:
            return JsonResponse({
                'success': False,
                'error': 'Введите пароль'
            }, status=400)
        
        if password != password_confirm:
            return JsonResponse({
                'success': False,
                'error': 'Пароли не совпадают'
            }, status=400)
        
        if len(password) < 6:
            return JsonResponse({
                'success': False,
                'error': 'Пароль должен содержать минимум 6 символов'
            }, status=400)
        
        # Проверка на существующего пользователя
        if User.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'error': 'Пользователь с таким email уже существует'
            }, status=400)
        
        # Создаем пользователя (username будет установлен в email через сигнал)
        user = User.objects.create_user(
            username=email,  # временно, сигнал потом обновит
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )
        
        # Если сигнал не сработал, устанавливаем вручную
        if user.username != email:
            user.username = email
            user.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Регистрация успешна! Теперь вы можете войти.',
            'user': {
                'id': user.id,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
        
@require_http_methods(["GET"])
@login_required
def get_user_history(request):
    """
    API для получения истории текущего пользователя
    """
    try:
        limit = int(request.GET.get('limit', 50))
        offset = int(request.GET.get('offset', 0))
        
        history = HistoryService.get_user_history(request.user, limit, offset)
        
        data = {
            'success': True,
            'history': [
                {
                    'id': h.id,
                    'description': h.description,
                    'datetime': h.datetime.isoformat(),
                    'datetime_formatted': h.datetime.strftime('%d.%m.%Y %H:%M:%S')
                }
                for h in history
            ],
            'total': History.objects.filter(user=request.user).count(),
            'limit': limit,
            'offset': offset
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
@staff_member_required
def get_all_history(request):
    """
    API для получения всей истории (только для персонала)
    """
    try:
        limit = int(request.GET.get('limit', 100))
        offset = int(request.GET.get('offset', 0))
        
        history = HistoryService.get_all_history(limit, offset)
        
        data = {
            'success': True,
            'history': [
                {
                    'id': h.id,
                    'user': h.user.username,
                    'user_id': h.user.id,
                    'description': h.description,
                    'datetime': h.datetime.isoformat(),
                    'datetime_formatted': h.datetime.strftime('%d.%m.%Y %H:%M:%S')
                }
                for h in history
            ],
            'total': History.objects.count(),
            'limit': limit,
            'offset': offset
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

@require_http_methods(["DELETE"])
@staff_member_required
def clear_old_history(request):
    """
    API для очистки старой истории (только для персонала)
    """
    try:
        days = int(request.GET.get('days', 30))
        deleted_count = HistoryService.delete_old_entries(days)
        
        # Логируем действие администратора
        HistoryService.add_entry(
            request.user,
            f"Администратор {request.user.username} удалил {deleted_count} старых записей истории (старше {days} дней)"
        )
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Удалено {deleted_count} записей старше {days} дней'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
            
# ========== АВТОРИЗАЦИЯ ==========

@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """Авторизация пользователя (по email или username)"""
    try:
        # ✅ ПРАВИЛЬНЫЙ способ получить данные - через request.data (DRF)
        # Не используем json.loads(request.body) напрямую!
        username_or_email = request.data.get('username', '').strip().lower()
        password = request.data.get('password', '')
        
        # Для отладки (можно убрать после проверки)
        print(f"Login attempt: {username_or_email}")
        
        if not username_or_email or not password:
            return JsonResponse({
                'success': False,
                'error': 'Введите логин/email и пароль'
            }, status=400)
        
        user = None
        
        # Сначала пробуем найти по email
        try:
            user_obj = User.objects.get(email=username_or_email)
            user = authenticate(request, username=user_obj.username, password=password)
        except User.DoesNotExist:
            # Если не нашли по email, пробуем по username
            user = authenticate(request, username=username_or_email, password=password)
        
        if user is not None:
            login(request, user)
            
            # Создаем или получаем токен
            token, created = Token.objects.get_or_create(user=user)
            
            response = JsonResponse({
                'success': True,
                'redirect_url': '/',
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'is_staff': user.is_staff,
                    'is_superuser': user.is_superuser,
                    'groups': [group.name for group in user.groups.all()]
                }
            })
            
            response.set_cookie(
                'auth_token',
                token.key,
                httponly=True,
                secure=False,
                samesite='Lax',
                max_age=timedelta(days=7)
            )
            
            return response
        else:
            return JsonResponse({
                'success': False,
                'error': 'Неверный логин/email или пароль'
            }, status=401)
            
    except Exception as e:
        print(f"Login error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    """Выход из системы"""
    try:
        # Удаляем токен
        token = request.COOKIES.get('auth_token')
        if token:
            Token.objects.filter(key=token).delete()
        
        logout(request)
        
        response = JsonResponse({'success': True, 'message': 'Выход выполнен'})
        response.delete_cookie('auth_token')
        response.delete_cookie('sessionid')
        response.delete_cookie('csrftoken')
        
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([AllowAny])
def check_auth(request):
    """Проверка авторизации"""
    token = request.COOKIES.get('auth_token')
    
    if token:
        try:
            token_obj = Token.objects.get(key=token)
            user = token_obj.user
            
            return JsonResponse({
                'authenticated': True,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'is_staff': user.is_staff,
                    'groups': [group.name for group in user.groups.all()]
                }
            })
        except Token.DoesNotExist:
            return JsonResponse({'authenticated': False})
    
    return JsonResponse({'authenticated': False})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_info(request):
    """Получение информации о текущем пользователе"""
    user = request.user
    return JsonResponse({
        'success': True,
        'user': {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'groups': [group.name for group in user.groups.all()],
            'permissions': list(user.get_all_permissions())
        }
    })
            

@login_required
def get_application_feedback(request, app_id):
    """
    Получение всех отзывов по заявке
    """
    try:
        application = Application.objects.get(id=app_id)
        
        # Проверка прав доступа
        if application.id_user != request.user and not user_can_view_all_applications(request.user):
            return JsonResponse({
                'success': False,
                'error': 'У вас нет прав для просмотра отзывов к этой заявке'
            }, status=403)
        
        feedbacks = Feedback.objects.filter(id_application=application).order_by('-created_at')
        
        return JsonResponse({
            'success': True,
            'application_id': application.id,
            'application_name': application.name,
            'feedbacks': [
                {
                    'id': f.id,
                    'name': f.name,
                    'user_id': f.id_user.id,
                    'user_name': f.id_user.username,
                    'comment': f.comment,
                    'created_at': f.created_at.isoformat(),
                    'created_at_formatted': f.created_at.strftime('%d.%m.%Y %H:%M:%S')
                }
                for f in feedbacks
            ],
            'total': feedbacks.count()
        })
        
    except Application.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Заявка не найдена'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_all_feedback(request):
    """
    Получение всех отзывов (для администраторов и пользователей с правами)
    """
    try:
        # Проверяем права
        if not user_can_view_all_applications(request.user):
            # Обычные пользователи видят только свои отзывы
            feedbacks = Feedback.objects.filter(id_user=request.user).order_by('-created_at')
        else:
            # Администраторы и менеджеры видят все отзывы
            feedbacks = Feedback.objects.all().order_by('-created_at')
        
        limit = int(request.GET.get('limit', 50))
        offset = int(request.GET.get('offset', 0))
        
        total = feedbacks.count()
        feedbacks_page = feedbacks[offset:offset+limit]
        
        return JsonResponse({
            'success': True,
            'feedbacks': [
                {
                    'id': f.id,
                    'name': f.name,
                    'user_id': f.id_user.id,
                    'user_name': f.id_user.username,
                    'user_email': f.id_user.email,
                    'application_id': f.id_application.id if f.id_application else None,
                    'application_name': f.id_application.name if f.id_application else None,
                    'comment': f.comment,
                    'created_at': f.created_at.isoformat(),
                    'created_at_formatted': f.created_at.strftime('%d.%m.%Y %H:%M:%S')
                }
                for f in feedbacks_page
            ],
            'total': total,
            'limit': limit,
            'offset': offset
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
                    
@login_required        
def get_rooms(request):
    """Получить список всех помещений с главным фото"""
    try:
        locations = Location.objects.all()
        rooms_list = []
        
        for location in locations:
            # Получаем главное фото для локации
            main_photo = Photo.objects.filter(id_location=location, is_main=True).first()
            if not main_photo:
                main_photo = Photo.objects.filter(id_location=location).first()
            
            photo_url = main_photo.photo.url if main_photo and main_photo.photo else None
            
            # Получаем оборудование в этой локации
            equipment_locations = EquipmentLocation.objects.filter(id_locations=location)
            equipment_count = equipment_locations.count()
            total_quantity = sum(eq.quantity for eq in equipment_locations)
            
            rooms_list.append({
                'id': location.id,
                'name': location.name,
                'size': location.size,
                'description': location.description,
                'photo_url': photo_url,
                'equipment_count': equipment_count,
                'total_quantity': total_quantity
            })
        
        return JsonResponse({
            'success': True,
            'rooms': rooms_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required 
def get_room_details(request):
    """Получить детали помещения с оборудованием и всеми фото"""
    try:
        room_id = request.GET.get('room_id')
        if not room_id:
            return JsonResponse({
                'success': False,
                'error': 'Не указан ID помещения'
            }, status=400)
        
        location = Location.objects.get(id=room_id)
        
        # Получаем все фото для галереи
        all_photos = Photo.objects.filter(id_location=location).order_by('-is_main', 'created_at')
        photos_list = []
        for photo in all_photos:
            photos_list.append({
                'id': photo.id,
                'name': photo.name,
                'url': photo.photo.url,
                'is_main': photo.is_main
            })
        
        # Получаем оборудование
        equipment_locations = EquipmentLocation.objects.filter(id_locations=location)
        equipment_list = []
        
        for eq_loc in equipment_locations:
            equipment_list.append({
                'name': eq_loc.id_equipments.name,
                'type': eq_loc.id_types_equipments.name,
                'quantity': eq_loc.quantity
            })
        
        return JsonResponse({
            'success': True,
            'room': {
                'id': location.id,
                'name': location.name,
                'size': location.size,
                'description': location.description,
                'photos': photos_list,
                'equipment_count': equipment_locations.count(),
                'total_quantity': sum(eq.quantity for eq in equipment_locations),
                'equipment': equipment_list
            }
        })
    except Location.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Помещение не найдено'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def export_orders_to_excel(request):
    """Экспорт выбранных заявок в Excel (с учетом прав просмотра)"""
    try:
        data = json.loads(request.body)
        order_ids = data.get('order_ids', [])
        
        print("=" * 60)
        print("ЭКСПОРТ ЗАЯВОК В EXCEL")
        print(f"ID заявок: {order_ids}")
        print(f"Пользователь: {request.user.username}")
        
        if not order_ids:
            return JsonResponse({
                'success': False,
                'error': 'Не выбраны заявки для экспорта'
            }, status=400)
        
        # Проверяем права пользователя
        from .views_helpers import user_can_view_all_applications
        
        if user_can_view_all_applications(request.user):
            # Пользователь может видеть все заявки
            applications = Application.objects.filter(id__in=order_ids).order_by('-created_at')
        else:
            # Обычный пользователь - только свои заявки
            applications = Application.objects.filter(id__in=order_ids, id_user=request.user).order_by('-created_at')
        
        if not applications.exists():
            return JsonResponse({
                'success': False,
                'error': 'Заявки не найдены или у вас нет прав для их просмотра'
            }, status=400)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заявки"
        
        # Стили (как в вашей существующей функции)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        location_fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
        slot_header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        slot_header_font = Font(bold=True, color="FFFFFF", size=10)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        for app in applications:
            orders = Order.objects.filter(id_application=app)
            
            # Заголовок заявки
            ws.merge_cells(f'A{current_row}:F{current_row}')
            title_cell = ws.cell(row=current_row, column=1, value=f"ЗАЯВКА №{app.id} - {app.name}")
            title_cell.font = Font(bold=True, size=14, color="1a56db")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # Информация о заявке
            ws.cell(row=current_row, column=1, value="Дата создания:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=app.created_at.strftime('%d.%m.%Y %H:%M') if app.created_at else "-")
            ws.cell(row=current_row, column=3, value="Статус:").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=dict(Application._meta.get_field('status').choices).get(app.status, app.status))
            
            # Добавляем информацию о пользователе (для админов)
            if user_can_view_all_applications(request.user):
                ws.cell(row=current_row, column=5, value="Пользователь:").font = Font(bold=True)
                ws.cell(row=current_row, column=6, value=app.id_user.username)
            current_row += 1
            
            if app.comment:
                ws.cell(row=current_row, column=1, value="Комментарий к заявке:").font = Font(bold=True)
                ws.merge_cells(f'B{current_row}:F{current_row}')
                ws.cell(row=current_row, column=2, value=app.comment)
                current_row += 1
            
            current_row += 1
            
            for order in orders:
                items = order.order_items.select_related(
                    'equipment_location__id_equipments',
                    'equipment_location__id_types_equipments',
                    'common_equipment_location__id_equipments',
                    'common_equipment_location__id_types_equipments'
                ).all()
                
                # Разделяем на обычные и слоты (как в вашей существующей функции)
                regular_items = []
                slot_groups = []
                
                for item in items:
                    if item.is_slot:
                        existing = None
                        for sg in slot_groups:
                            if sg['start'] == item.slot_date_start and sg['end'] == item.slot_date_end:
                                existing = sg
                                break
                        if existing:
                            existing['items'].append(item)
                        else:
                            slot_groups.append({
                                'start': item.slot_date_start,
                                'end': item.slot_date_end,
                                'items': [item]
                            })
                    else:
                        regular_items.append(item)
                
                slot_groups.sort(key=lambda x: x['start'])
                
                # Собираем все типы оборудования
                all_types = set()
                for item in regular_items:
                    if item.equipment_location:
                        all_types.add(item.equipment_location.id_types_equipments.name)
                    else:
                        all_types.add(item.common_equipment_location.id_types_equipments.name)
                
                for slot_group in slot_groups:
                    for item in slot_group['items']:
                        if item.equipment_location:
                            all_types.add(item.equipment_location.id_types_equipments.name)
                        else:
                            all_types.add(item.common_equipment_location.id_types_equipments.name)
                
                types_list = sorted(all_types)
                
                if not types_list:
                    continue
                
                total_cols = len(types_list) + 1
                last_col_letter = openpyxl.utils.get_column_letter(total_cols)
                
                # Дата и время заказа
                date_start_str = order.date_time_start.strftime('%d.%m.%Y %H:%M') if order.date_time_start else "-"
                date_end_str = order.date_time_end.strftime('%d.%m.%Y %H:%M') if order.date_time_end else "Не завершен"
                date_range = f"📅 {date_start_str} - {date_end_str}"
                
                ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
                date_cell = ws.cell(row=current_row, column=1, value=date_range)
                date_cell.font = Font(bold=True, size=10, color="666666")
                date_cell.alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1
                
                if order.comment:
                    ws.cell(row=current_row, column=1, value="💬 Комментарий к заказу:").font = Font(bold=True)
                    ws.merge_cells(f'B{current_row}:{last_col_letter}{current_row}')
                    ws.cell(row=current_row, column=2, value=order.comment)
                    current_row += 1
                
                # Заголовки таблицы
                location_cell = ws.cell(row=current_row, column=1, value="ЛОКАЦИЯ")
                location_cell.font = header_font
                location_cell.fill = header_fill
                location_cell.alignment = header_alignment
                location_cell.border = thin_border
                
                for col_idx, type_name in enumerate(types_list, start=2):
                    cell = ws.cell(row=current_row, column=col_idx, value=type_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                current_row += 1
                
                location_name = order.id_location.name if order.id_location else "Не указана"
                table_start_row = current_row
                
                # Функция для преобразования оборудования в матрицу
                def equipment_to_matrix(equipment_items):
                    if not equipment_items:
                        return []
                    
                    types_data = {}
                    for item in equipment_items:
                        if item.equipment_location:
                            type_name = item.equipment_location.id_types_equipments.name
                            equipment_name = item.equipment_location.id_equipments.name
                        else:
                            type_name = item.common_equipment_location.id_types_equipments.name
                            equipment_name = f"🌍 {item.common_equipment_location.id_equipments.name}"
                        
                        if type_name not in types_data:
                            types_data[type_name] = []
                        types_data[type_name].append({
                            'name': equipment_name,
                            'quantity': item.quantity
                        })
                    
                    for t_name in types_data:
                        types_data[t_name].sort(key=lambda x: x['name'])
                    
                    max_rows = max((len(types_data.get(t, [])) for t in types_list), default=0)
                    
                    rows_data = []
                    for row_idx in range(max_rows):
                        row_cells = []
                        for type_name in types_list:
                            equipment_list = types_data.get(type_name, [])
                            if row_idx < len(equipment_list):
                                eq = equipment_list[row_idx]
                                row_cells.append(f"{eq['name']} - {eq['quantity']} шт.")
                            else:
                                row_cells.append("—")
                        rows_data.append(row_cells)
                    
                    return rows_data
                
                # Собираем все строки таблицы
                all_table_rows = []
                
                # Обычные позиции
                regular_matrix = equipment_to_matrix(regular_items)
                for row_cells in regular_matrix:
                    all_table_rows.append({'type': 'equipment', 'cells': row_cells})
                
                # Слоты
                for slot_group in slot_groups:
                    slot_start_str = slot_group['start'].strftime('%H:%M') if slot_group['start'] else "-"
                    slot_end_str = slot_group['end'].strftime('%H:%M') if slot_group['end'] else "-"
                    slot_text = f"⏰ СЛОТ: {slot_start_str} - {slot_end_str}"
                    all_table_rows.append({'type': 'slot_header', 'text': slot_text})
                    
                    slot_matrix = equipment_to_matrix(slot_group['items'])
                    for row_cells in slot_matrix:
                        all_table_rows.append({'type': 'equipment', 'cells': row_cells})
                
                # Вывод всех строк
                for row_data in all_table_rows:
                    if row_data['type'] == 'slot_header':
                        empty_cell = ws.cell(row=current_row, column=1, value="")
                        empty_cell.border = thin_border
                        empty_cell.fill = location_fill
                        
                        ws.merge_cells(f'B{current_row}:{last_col_letter}{current_row}')
                        slot_cell = ws.cell(row=current_row, column=2, value=row_data['text'])
                        slot_cell.font = slot_header_font
                        slot_cell.fill = slot_header_fill
                        slot_cell.alignment = Alignment(horizontal="center", vertical="center")
                        slot_cell.border = thin_border
                        current_row += 1
                    else:
                        empty_cell = ws.cell(row=current_row, column=1, value="")
                        empty_cell.border = thin_border
                        empty_cell.fill = location_fill
                        
                        for col_idx, cell_value in enumerate(row_data['cells'], start=2):
                            if cell_value == "—":
                                cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                                cell.alignment = center_alignment
                                cell.border = thin_border
                            else:
                                cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                                cell.alignment = left_alignment
                                cell.border = thin_border
                        
                        current_row += 1
                
                # Заполняем первую строку локацией
                first_data_row = table_start_row
                if current_row > table_start_row:
                    location_cell = ws.cell(row=first_data_row, column=1, value=location_name)
                    location_cell.alignment = left_alignment
                    location_cell.border = thin_border
                    location_cell.fill = location_fill
                    location_cell.font = Font(bold=True)
                    
                    if current_row > table_start_row + 1:
                        ws.merge_cells(
                            start_row=table_start_row,
                            start_column=1,
                            end_row=current_row - 1,
                            end_column=1
                        )
                        merged_cell = ws.cell(row=table_start_row, column=1)
                        merged_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                current_row += 2
            
            current_row += 1
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 35
        
        # Определяем максимальное количество типов для ширины колонок
        max_types_count = 0
        for app in applications:
            for order in Order.objects.filter(id_application=app):
                items = order.order_items.all()
                types_set = set()
                for item in items:
                    if item.equipment_location:
                        types_set.add(item.equipment_location.id_types_equipments.name)
                    else:
                        types_set.add(item.common_equipment_location.id_types_equipments.name)
                max_types_count = max(max_types_count, len(types_set))
        
        for col_idx in range(2, max_types_count + 3):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 35
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="applications_export.xlsx"'
        
        return response
        
    except Exception as e:
        print(f"Ошибка экспорта: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        

@csrf_exempt
@require_http_methods(["GET"])
def check_equipment_location(request):
    """Проверка существования EquipmentLocation"""
    location_id = request.GET.get('location_id')
    equipment_id = request.GET.get('equipment_id')
    
    if not location_id or not equipment_id:
        return JsonResponse({'exists': False, 'error': 'Missing parameters'})
    
    try:
        exists = EquipmentLocation.objects.filter(
            id_locations_id=location_id,
            id_equipments_id=equipment_id
        ).exists()
        
        # Если не существует, попробуем найти все доступные для отладки
        if not exists:
            available = EquipmentLocation.objects.filter(
                id_locations_id=location_id
            ).values_list('id_equipments_id', flat=True)
            
            return JsonResponse({
                'exists': False,
                'available_equipment': list(available),
                'location_id': location_id,
                'equipment_id': equipment_id
            })
        
        return JsonResponse({'exists': True})
    except Exception as e:
        return JsonResponse({'exists': False, 'error': str(e)})
    

@login_required    
@csrf_exempt
@require_http_methods(["POST"])
def save_multiple_orders(request):
    """Сохранение нескольких заказов"""
    try:
        data = json.loads(request.body)
        orders_data = data.get('orders', [])
        
        if not orders_data:
            return JsonResponse({
                'success': False,
                'error': 'Нет данных для сохранения'
            }, status=400)
        
        results = []
        for order_data in orders_data:
            try:
                # Создаем заказ
                order = Order.objects.create(
                    date_time_start=order_data.get('date_time_start'),
                    date_time_end=order_data.get('date_time_end') if order_data.get('date_time_end') else None
                )
                
                items_added = 0
                for item in order_data.get('items', []):
                    try:
                        equipment_location = EquipmentLocation.objects.get(
                            id_locations_id=item['location_id'],
                            id_equipments_id=item['equipment_id']
                        )
                        
                        quantity = min(item['quantity'], equipment_location.quantity)
                        
                        if quantity > 0:
                            OrderItem.objects.create(
                                order=order,
                                equipment_location=equipment_location,
                                quantity=quantity
                            )
                            items_added += 1
                    except EquipmentLocation.DoesNotExist:
                        continue
                
                if items_added > 0:
                    results.append({
                        'success': True,
                        'order_id': order.id,
                        'items_added': items_added
                    })
                else:
                    order.delete()
                    results.append({
                        'success': False,
                        'error': 'Нет валидных позиций'
                    })
            except Exception as e:
                results.append({
                    'success': False,
                    'error': str(e)
                })
        
        successful = [r for r in results if r.get('success')]
        failed = [r for r in results if not r.get('success')]
        
        return JsonResponse({
            'success': len(successful) > 0,
            'total': len(orders_data),
            'successful': len(successful),
            'failed': len(failed),
            'results': results,
            'message': f'Создано заказов: {len(successful)} из {len(orders_data)}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        

@login_required 
@csrf_exempt
@require_http_methods(["POST"])
def save_order(request):
    """Сохранение заказа в базу данных (с привязкой к заявке и локации)"""
    try:
        data = json.loads(request.body)
        
        print("=" * 60)
        print("ПОЛУЧЕН ЗАПРОС НА СОХРАНЕНИЕ ЗАКАЗА")
        print(f"Пользователь: {request.user.username} (ID: {request.user.id})")
        print(f"Дата начала заказа: {data.get('date_time_start')}")
        print(f"Дата окончания заказа: {data.get('date_time_end')}")
        print(f"Дата начала заявки (из date-section): {data.get('application_date_start')}")
        print(f"Дата окончания заявки (из date-section): {data.get('application_date_end')}")
        print(f"Application ID: {data.get('application_id')}")
        print(f"Location ID: {data.get('location_id')}")
        print(f"Тип заказа: {data.get('type')}")  # 'slots' или 'regular'
        print(f"Количество слотов: {len(data.get('slots', []))}")
        
        # ========== ПРОВЕРКА: ЭТО ЗАКАЗ СО СЛОТАМИ ==========
        is_slots_order = data.get('type') == 'slots'
        slots_data = data.get('slots', [])
        
        if not data.get('date_time_start'):
            return JsonResponse({
                'success': False,
                'error': 'Не указана дата начала заказа'
            }, status=400)
        
        items_data = data.get('items', [])
        
        # Для слотов items_data может быть пустым, оборудование внутри слотов
        if not items_data and not slots_data:
            return JsonResponse({
                'success': False,
                'error': 'Нет позиций для сохранения'
            }, status=400)
        
        application_id = data.get('application_id')
        application = None
        is_new_application = False
        
        app_date_start = data.get('application_date_start')
        app_date_end = data.get('application_date_end')
        
        print(f"Даты для Application: начало={app_date_start}, окончание={app_date_end}")
        
        # ========== СОБИРАЕМ ВСЕ ТИПЫ ОБОРУДОВАНИЯ ИЗ ТЕКУЩЕГО ЗАКАЗА ==========
        current_equipment_types = set()
        
        # Если это слоты, собираем типы из слотов
        if is_slots_order and slots_data:
            for slot in slots_data:
                for item in slot.get('equipment', []):
                    equipment_id = item.get('equipment_id')
                    is_common = item.get('is_common', False)
                    
                    if is_common:
                        common_equipment = CommonEquipmentLocation.objects.filter(
                            id_equipments_id=equipment_id
                        ).first()
                        if common_equipment:
                            current_equipment_types.add(common_equipment.id_types_equipments.id)
                    else:
                        equipment_location = EquipmentLocation.objects.filter(
                            id_equipments_id=equipment_id
                        ).first()
                        if equipment_location:
                            current_equipment_types.add(equipment_location.id_types_equipments.id)
        else:
            # Обычный режим
            for item in items_data:
                equipment_id = item.get('equipment_id')
                is_common = item.get('is_common', False)
                
                if is_common:
                    common_equipment = CommonEquipmentLocation.objects.filter(
                        id_equipments_id=equipment_id
                    ).first()
                    if common_equipment:
                        current_equipment_types.add(common_equipment.id_types_equipments.id)
                else:
                    equipment_location = EquipmentLocation.objects.filter(
                        id_equipments_id=equipment_id
                    ).first()
                    if equipment_location:
                        current_equipment_types.add(equipment_location.id_types_equipments.id)
        
        print(f"Типы оборудования: {current_equipment_types}")
        
        # ========== СОЗДАНИЕ ИЛИ ПОЛУЧЕНИЕ ЗАЯВКИ ==========
        if application_id:
            try:
                application = Application.objects.get(id=application_id, id_user=request.user)
                print(f"Найдена существующая заявка №{application.id}")
                
                if app_date_start:
                    application.date_start = app_date_start
                if app_date_end:
                    application.date_end = app_date_end
                application.save(update_fields=['date_start', 'date_end'])
                
                # Обновляем согласования
                all_equipment_types = get_all_equipment_types_for_application(application)
                all_equipment_types.update(current_equipment_types)
                update_application_approvals(application, all_equipment_types)
                
            except Application.DoesNotExist:
                print(f"Заявка с ID {application_id} не найдена")
                return JsonResponse({
                    'success': False,
                    'error': 'Заявка не найдена'
                }, status=404)
        else:
            application_name = data.get('application_name')
            if not application_name:
                return JsonResponse({
                    'success': False,
                    'error': 'Не указано название заявки'
                }, status=400)
            
            application_comment = data.get('application_comment', '')
            
            application = Application.objects.create(
                name=application_name,
                id_user=request.user,
                comment=application_comment,
                status='new',
                date_start=app_date_start,
                date_end=app_date_end
            )
            is_new_application = True
            print(f"Создана новая заявка №{application.id}")
            
            create_application_approvals(application, current_equipment_types)
            
            # Отправляем уведомление
            try:
                from .email_utils import EmailService
                EmailService.send_new_application_notification(application, request.user)
                print(f"Уведомление о новой заявке #{application.id} отправлено")
            except Exception as e:
                print(f"Ошибка отправки уведомления: {e}")
        
        # ========== ПОЛУЧАЕМ ЛОКАЦИЮ ==========
        location_id = data.get('location_id')
        location = None
        if location_id:
            try:
                location = Location.objects.get(id=location_id)
                print(f"Найдена локация: {location.name}")
            except Location.DoesNotExist:
                print(f"Локация с ID {location_id} не найдена")
        
        # ========== СОЗДАНИЕ ЗАКАЗА ==========
        date_start = data.get('date_time_start')
        date_end = data.get('date_time_end')
        
        order = Order.objects.create(
            id_user=request.user,
            id_application=application,
            id_location=location,
            date_time_start=date_start,
            date_time_end=date_end if date_end else None,
            comment=data.get('comment', '')
        )
        
        print(f"Создан заказ №{order.id} для локации {location.name if location else 'Не указана'}")
        
        items_added = 0
        items_failed = 0
        
        # ========== ОБРАБОТКА СЛОТОВ ==========
        if is_slots_order and slots_data:
            print(f"\n=== ОБРАБОТКА СЛОТОВ ({len(slots_data)} слотов) ===")
            
            for slot_idx, slot in enumerate(slots_data):
                slot_date_start = slot.get('date_start')
                slot_date_end = slot.get('date_end')
                slot_equipment = slot.get('equipment', [])
                
                print(f"\nСлот {slot_idx + 1}: {slot_date_start} - {slot_date_end}")
                print(f"  Оборудование: {len(slot_equipment)} позиций")
                
                for eq_idx, eq in enumerate(slot_equipment):
                    try:
                        equipment_id = eq.get('equipment_id')
                        quantity = eq.get('quantity', 0)
                        is_common = eq.get('is_common', False)
                        
                        print(f"    Оборудование ID: {equipment_id}, is_common={is_common}, quantity={quantity}")
                        
                        if quantity <= 0:
                            print(f"    ✗ Количество = 0, пропуск")
                            continue
                        
                        if is_common:
                            # Общее оборудование
                            common_equipment = CommonEquipmentLocation.objects.filter(
                                id_equipments_id=equipment_id
                            ).first()
                            
                            if not common_equipment:
                                print(f"    ✗ Общее оборудование ID {equipment_id} не найдено")
                                items_failed += 1
                                continue
                            
                            # Проверка доступности для общего оборудования
                            busy_quantity = OrderItem.objects.filter(
                                common_equipment_location=common_equipment,
                                order__date_time_start__lt=slot_date_end,
                                order__date_time_end__gt=slot_date_start
                            ).exclude(
                                order__date_time_end__isnull=True
                            ).exclude(
                                order__id_application__status='cancelled'
                            ).exclude(
                                order=order
                            ).aggregate(total=Sum('quantity'))['total'] or 0
                            
                            active_busy = OrderItem.objects.filter(
                                common_equipment_location=common_equipment,
                                order__date_time_end__isnull=True,
                                order__date_time_start__lt=slot_date_end
                            ).exclude(
                                order__id_application__status='cancelled'
                            ).exclude(
                                order=order
                            ).aggregate(total=Sum('quantity'))['total'] or 0
                            
                            total_busy = busy_quantity + active_busy
                            available = common_equipment.quantity - total_busy
                            final_quantity = min(quantity, max(available, 0))
                            
                            if final_quantity > 0:
                                OrderItem.objects.create(
                                    order=order,
                                    common_equipment_location=common_equipment,
                                    quantity=final_quantity,
                                    is_slot=True,
                                    slot_date_start=slot_date_start,
                                    slot_date_end=slot_date_end
                                )
                                items_added += 1
                                print(f"    ✓ Общее: {common_equipment.id_equipments.name} x{final_quantity} [СЛОТ]")
                            else:
                                print(f"    ✗ Общее: {common_equipment.id_equipments.name} - недостаточно (доступно {available})")
                                items_failed += 1
                        
                        else:
                            # Обычное оборудование из локации
                            if not location_id:
                                print(f"    ✗ Не указана локация для обычного оборудования")
                                items_failed += 1
                                continue
                            
                            # ВАЖНО: equipment_id из фронтенда - это id из Equipment (оборудования)
                            # Нужно найти EquipmentLocation по id_equipments и id_locations
                            equipment_location = EquipmentLocation.objects.filter(
                                id_locations_id=location_id,
                                id_equipments_id=equipment_id  # equipment_id - это id из таблицы Equipment
                            ).first()
                            
                            if not equipment_location:
                                # Пробуем найти по другому полю (если equipment_id - это id из EquipmentLocation)
                                equipment_location = EquipmentLocation.objects.filter(
                                    id=equipment_id
                                ).first()
                                
                                if equipment_location and equipment_location.id_locations_id != location_id:
                                    print(f"    ✗ EquipmentLocation найден, но для другой локации")
                                    equipment_location = None
                            
                            if not equipment_location:
                                print(f"    ✗ EquipmentLocation не найден (location={location_id}, equipment_id={equipment_id})")
                                # Выводим доступное оборудование для отладки
                                available_eq = EquipmentLocation.objects.filter(
                                    id_locations_id=location_id
                                ).values_list('id_equipments_id', flat=True)[:5]
                                print(f"    Доступные equipment_id в этой локации: {list(available_eq)}")
                                items_failed += 1
                                continue
                            
                            final_quantity = min(quantity, equipment_location.quantity)
                            
                            if final_quantity > 0:
                                OrderItem.objects.create(
                                    order=order,
                                    equipment_location=equipment_location,
                                    quantity=final_quantity,
                                    is_slot=True,
                                    slot_date_start=slot_date_start,
                                    slot_date_end=slot_date_end
                                )
                                items_added += 1
                                print(f"    ✓ {equipment_location.id_equipments.name} x{final_quantity} [СЛОТ]")
                            else:
                                print(f"    ✗ {equipment_location.id_equipments.name} - недостаточно (доступно {equipment_location.quantity})")
                                items_failed += 1
                                
                    except Exception as e:
                        items_failed += 1
                        print(f"    ✗ Ошибка при сохранении оборудования: {str(e)}")
                        import traceback
                        traceback.print_exc()
                
        # ========== ОБЫЧНЫЙ РЕЖИМ (НЕ СЛОТЫ) ==========
        else:
            print("\n=== ОБЫЧНЫЙ РЕЖИМ ===")
            for idx, item in enumerate(items_data):
                try:
                    is_common = item.get('is_common', False)
                    equipment_id = item.get('equipment_id')
                    requested_qty = item.get('quantity', 0)
                    item_location_id = item.get('location_id')
                    
                    print(f"  Позиция {idx + 1}: equipment_id={equipment_id}, location_id={item_location_id}, is_common={is_common}, qty={requested_qty}")
                    
                    if is_common:
                        common_equipment = CommonEquipmentLocation.objects.filter(
                            id_equipments_id=equipment_id
                        ).first()
                        
                        if not common_equipment:
                            items_failed += 1
                            print(f"    ✗ Общее оборудование не найдено")
                            continue
                        
                        busy_quantity = OrderItem.objects.filter(
                            common_equipment_location=common_equipment,
                            order__date_time_start__lt=date_end,
                            order__date_time_end__gt=date_start
                        ).exclude(
                            order__date_time_end__isnull=True
                        ).exclude(
                            order__id_application__status='cancelled'
                        ).exclude(
                            order=order
                        ).aggregate(total=Sum('quantity'))['total'] or 0
                        
                        active_busy = OrderItem.objects.filter(
                            common_equipment_location=common_equipment,
                            order__date_time_end__isnull=True,
                            order__date_time_start__lt=date_end
                        ).exclude(
                            order__id_application__status='cancelled'
                        ).exclude(
                            order=order
                        ).aggregate(total=Sum('quantity'))['total'] or 0
                        
                        total_busy = busy_quantity + active_busy
                        available = common_equipment.quantity - total_busy
                        quantity = min(requested_qty, max(available, 0))
                        
                        if quantity > 0:
                            OrderItem.objects.create(
                                order=order,
                                common_equipment_location=common_equipment,
                                quantity=quantity
                            )
                            items_added += 1
                            print(f"    ✓ Сохранено общее оборудование: {common_equipment.id_equipments.name} x{quantity}")
                        else:
                            items_failed += 1
                            
                    else:
                        if not item_location_id:
                            items_failed += 1
                            print(f"    ✗ Не указан location_id для обычного оборудования")
                            continue
                        
                        equipment_location = EquipmentLocation.objects.filter(
                            id_locations_id=item_location_id,
                            id_equipments_id=equipment_id
                        ).first()
                        
                        if not equipment_location:
                            items_failed += 1
                            print(f"    ✗ EquipmentLocation не найден")
                            continue
                        
                        total_quantity = equipment_location.quantity
                        quantity = min(requested_qty, total_quantity)
                        
                        if quantity > 0:
                            OrderItem.objects.create(
                                order=order,
                                equipment_location=equipment_location,
                                quantity=quantity
                            )
                            items_added += 1
                            print(f"    ✓ Сохранено: {equipment_location.id_equipments.name} x{quantity}")
                        else:
                            items_failed += 1
                        
                except Exception as e:
                    items_failed += 1
                    print(f"    ✗ Ошибка: {str(e)}")
                    continue
        
        print(f"\nИТОГ: добавлено {items_added} позиций, пропущено {items_failed}")
        print("=" * 60)
        
        if items_added == 0:
            order.delete()
            if is_new_application:
                application.delete()
            return JsonResponse({
                'success': False,
                'error': 'Не удалось добавить ни одной позиции'
            }, status=400)
        
        HistoryService.add_entry(
            request.user,
            f"Создан заказ #{order.id} для локации {location.name if location else 'Общее оборудование'} с {items_added} позициями"
        )
        
        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'application_id': application.id if application else None,
            'location_id': location.id if location else None,
            'items_added': items_added,
            'items_failed': items_failed,
            'is_slots_order': is_slots_order,
            'slots_count': len(slots_data) if is_slots_order else 0,
            'message': f'Заказ №{order.id} сохранен. Добавлено позиций: {items_added}'
        })
        
    except Exception as e:
        print(f"КРИТИЧЕСКАЯ ОШИБКА: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        

def index(request):
    """Главная страница - проверяем авторизацию"""
    # Если пользователь не авторизован, возвращаем страницу с формой входа
    if not request.user.is_authenticated:
        locations = Location.objects.all()
        type_equipments = TypeEquipment.objects.all()
        context = {
            'locations': locations,
            'type_equipments': type_equipments,
        }
        return render(request, 'mainapp/index.html', context)
    
    # Авторизованный пользователь видит полную страницу
    locations = Location.objects.all()
    type_equipments = TypeEquipment.objects.all()
    context = {
        'locations': locations,
        'type_equipments': type_equipments,
    }
    return render(request, 'mainapp/index.html', context)


@login_required
def get_location_photo(request, location_id):
    """Получить главное фото локации"""
    try:
        # Сначала пытаемся найти главное фото
        photo = Photo.objects.filter(id_location_id=location_id, is_main=True).first()
        # Если главного нет, берем любое первое
        if not photo:
            photo = Photo.objects.filter(id_location_id=location_id).first()
        
        if photo and photo.photo:
            return JsonResponse({
                'success': True,
                'photo_url': photo.photo.url,
                'photo_name': photo.name,
                'is_main': photo.is_main
            })
        return JsonResponse({'success': False, 'error': 'Фото не найдено'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_equipment_by_location(request):
    """Получить оборудование по локации и типу"""
    location_id = request.GET.get('location_id')
    type_ids = request.GET.getlist('type_ids[]')
    
    if not location_id:
        return JsonResponse({'success': False, 'error': 'Не выбрана локация'})
    
    try:
        equipment_locations = EquipmentLocation.objects.filter(
            id_locations_id=location_id
        )
        
        if type_ids:
            equipment_locations = equipment_locations.filter(
                id_types_equipments_id__in=type_ids
            )
        
        equipment_list = []
        for eq_loc in equipment_locations:
            equipment_list.append({
                'id': eq_loc.id,  # ID из EquipmentLocation
                'equipment_id': eq_loc.id_equipments.id,  # ВАЖНО: ID из Equipment
                'name': eq_loc.id_equipments.name,
                'type_id': eq_loc.id_types_equipments_id,
                'type_name': eq_loc.id_types_equipments.name,
                'quantity': eq_loc.quantity,
                'max_quantity': eq_loc.quantity,
            })
        
        return JsonResponse({
            'success': True,
            'equipment': equipment_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_common_equipment(request):
    """Получить общее оборудование"""
    try:
        type_ids = request.GET.getlist('type_ids[]')
        
        common_equipment = CommonEquipmentLocation.objects.all()
        
        if type_ids:
            common_equipment = common_equipment.filter(
                id_types_equipments_id__in=type_ids
            )
        
        equipment_list = []
        for eq in common_equipment:
            equipment_list.append({
                'id': eq.id,
                'equipment_id': eq.id_equipments.id,
                'name': eq.id_equipments.name,
                'type_id': eq.id_types_equipments_id,
                'type_name': eq.id_types_equipments.name,
                'quantity': eq.quantity,
                'max_quantity': eq.quantity,
                'is_common': True  # Важно: флаг общего оборудования
            })
        
        return JsonResponse({
            'success': True,
            'equipment': equipment_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def export_to_excel(request):
    """Экспорт данных в Excel"""
    try:
        data = json.loads(request.body)
        selected_equipment = data.get('equipment', [])
        location_name = data.get('location_name', '')
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Оборудование"
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовки
        headers = ['Наименование оборудования', 'Тип', 'Количество', 'Примечание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Заполняем данные
        for row, item in enumerate(selected_equipment, 2):
            ws.cell(row=row, column=1, value=item.get('name', ''))
            ws.cell(row=row, column=2, value=item.get('type_name', ''))
            ws.cell(row=row, column=3, value=item.get('selected_quantity', 0))
            ws.cell(row=row, column=4, value='')
        
        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Добавляем информацию о локации
        ws.insert_rows(1)
        ws.merge_cells('A1:D1')
        cell = ws.cell(row=1, column=1, value=f"Локация: {location_name}")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        
        # Сохраняем в буфер
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Отправляем файл
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="equipment_{location_name}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
 

def user_can_view_all_applications(user):
    """Проверка, может ли пользователь видеть все заявки"""
    return user.is_superuser or user.groups.filter(name='ViewApplications').exists() or user.groups.filter(name='EditApplications').exists()

def user_can_edit_all_applications(user):
    """Проверка, может ли пользователь редактировать все заявки"""
    return user.is_superuser or user.groups.filter(name='EditApplications').exists()


@login_required
def get_orders(request):
    """Получить список заявок с учетом прав пользователя"""
    try:
        if user_can_view_all_applications(request.user):
            applications = Application.objects.all().order_by('-created_at')
        else:
            applications = Application.objects.filter(id_user=request.user).order_by('-created_at')
        
        apps_list = []
        
        for app in applications:
            orders = Order.objects.filter(id_application=app)
            first_order = orders.first()
            
            app_date_start = app.date_start
            app_date_end = app.date_end
            
            if not app_date_start and first_order:
                app_date_start = first_order.date_time_start
            if not app_date_end and first_order:
                app_date_end = first_order.date_time_end
            
            status_display = dict(Application._meta.get_field('status').choices).get(app.status, app.status)
            is_active = app.status == 'new' or app.status == 'in_progress'
            
            can_edit = (app.id_user == request.user) or user_can_edit_all_applications(request.user)
            
            # Получаем список согласований для заявки
            approvals = ApplicationApproval.objects.filter(id_application=app).select_related('id_department', 'id_user')
            
            # Получаем все позиции оборудования из всех заказов заявки
            all_order_items = []
            for order in orders:
                items = OrderItem.objects.filter(order=order).select_related(
                    'equipment_location__id_equipments',
                    'equipment_location__id_types_equipments',
                    'common_equipment_location__id_equipments',
                    'common_equipment_location__id_types_equipments'
                )
                all_order_items.extend(items)
            
            # Группируем оборудование по подразделениям
            department_equipment_status = {}
            
            for item in all_order_items:
                # Получаем тип оборудования
                if item.equipment_location:
                    equipment_type_id = item.equipment_location.id_types_equipments.id
                    equipment_type_name = item.equipment_location.id_types_equipments.name
                elif item.common_equipment_location:
                    equipment_type_id = item.common_equipment_location.id_types_equipments.id
                    equipment_type_name = item.common_equipment_location.id_types_equipments.name
                else:
                    continue
                
                # Находим подразделения, связанные с этим типом оборудования
                departments_for_type = Department.objects.filter(
                    department_type_equipment__id_type_equipment_id=equipment_type_id
                )
                
                for dept in departments_for_type:
                    if dept.id not in department_equipment_status:
                        department_equipment_status[dept.id] = {
                            'department_name': dept.name,
                            'total_items': 0,
                            'agreed_items': 0,
                            'all_agreed': False
                        }
                    
                    department_equipment_status[dept.id]['total_items'] += 1
                    if item.is_agreed:
                        department_equipment_status[dept.id]['agreed_items'] += 1
            
            # Определяем, все ли оборудование согласовано для каждого подразделения
            for dept_id in department_equipment_status:
                status = department_equipment_status[dept_id]
                status['all_agreed'] = (status['total_items'] > 0 and status['agreed_items'] == status['total_items'])
            
            approval_statuses = []
            for approval in approvals:
                dept_status = department_equipment_status.get(approval.id_department.id, {})
                all_equipment_agreed = dept_status.get('all_agreed', False)
                
                approval_statuses.append({
                    'department_id': approval.id_department.id,
                    'department_name': approval.id_department.name,
                    'is_agreed': approval.is_agreed,
                    'status_text': '✅ Согласовано' if approval.is_agreed else ' На согласовании ',
                    'status_class': 'approved' if approval.is_agreed else 'pending',
                    'user_name': approval.id_user.username if approval.id_user else None,
                    'date_agreed': approval.date_agreed.isoformat() if approval.date_agreed else None,
                    'all_equipment_agreed': all_equipment_agreed,  # Флаг, что всё оборудование подразделения согласовано
                    'total_items': dept_status.get('total_items', 0),
                    'agreed_items': dept_status.get('agreed_items', 0)
                })
            
            apps_list.append({
                'id': app.id,
                'application_name': app.name,
                'date_time_start': app_date_start.isoformat() if app_date_start else None,
                'date_time_end': app_date_end.isoformat() if app_date_end else None,
                'application_date_start': app.date_start.isoformat() if app.date_start else None,
                'application_date_end': app.date_end.isoformat() if app.date_end else None,
                'is_active': is_active,
                'status': app.status,
                'status_display': status_display,
                'comment': app.comment or '',
                'total_orders': orders.count(),
                'total_quantity': sum(order.total_quantity for order in orders),
                'can_edit': can_edit,
                'user_id': app.id_user.id,
                'user_name': app.id_user.username,
                'approvals': approval_statuses
            })
        
        return JsonResponse({
            'success': True,
            'orders': apps_list,
            'user_permissions': {
                'can_view_all': user_can_view_all_applications(request.user),
                'can_edit_all': user_can_edit_all_applications(request.user),
                'is_superuser': request.user.is_superuser
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def get_order_items(request):
    """Получить позиции заявки с проверкой прав"""
    try:
        application_id = request.GET.get('order_id')
        if not application_id:
            return JsonResponse({
                'success': False,
                'error': 'Не указан ID заявки'
            }, status=400)
        
        application = Application.objects.get(id=application_id)
        
        # Проверка прав доступа к заявке
        if application.id_user != request.user and not user_can_view_all_applications(request.user):
            return JsonResponse({
                'success': False,
                'error': 'У вас нет прав для просмотра этой заявки'
            }, status=403)
        
        orders = Order.objects.filter(id_application=application)
        
        items_list = []
        for order in orders:
            order_items = order.order_items.all()
            for item in order_items:
                if item.equipment_location:
                    items_list.append({
                        'id': item.id,
                        'order_id': order.id,
                        'location_id': item.equipment_location.id_locations.id,
                        'location_name': item.equipment_location.id_locations.name,
                        'equipment_id': item.equipment_location.id_equipments.id,
                        'equipment_name': item.equipment_location.id_equipments.name,
                        'type_name': item.equipment_location.id_types_equipments.name,
                        'quantity': item.quantity,
                        'is_common': False,
                        'date_start': order.date_time_start.isoformat(),
                        'date_end': order.date_time_end.isoformat() if order.date_time_end else None,
                        'order_comment': order.comment or '',
                        'can_provide': item.can_provide,
                        'is_agreed': item.is_agreed,
                        'is_slot': item.is_slot,
                        'slot_date_start': item.slot_date_start.isoformat() if item.slot_date_start else None,
                        'slot_date_end': item.slot_date_end.isoformat() if item.slot_date_end else None
                    })
                elif item.common_equipment_location:
                    items_list.append({
                        'id': item.id,
                        'order_id': order.id,
                        'location_id': order.id_location.id if order.id_location else None,
                        'location_name': order.id_location.name if order.id_location else 'Не указана',
                        'equipment_id': item.common_equipment_location.id_equipments.id,
                        'equipment_name': item.common_equipment_location.id_equipments.name,
                        'type_name': item.common_equipment_location.id_types_equipments.name,
                        'quantity': item.quantity,
                        'is_common': True,
                        'date_start': order.date_time_start.isoformat(),
                        'date_end': order.date_time_end.isoformat() if order.date_time_end else None,
                        'order_comment': order.comment or '',
                        'can_provide': item.can_provide,
                        'is_agreed': item.is_agreed,
                        'is_slot': item.is_slot,
                        'slot_date_start': item.slot_date_start.isoformat() if item.slot_date_start else None,
                        'slot_date_end': item.slot_date_end.isoformat() if item.slot_date_end else None
                    })
        
        return JsonResponse({
            'success': True,
            'items': items_list,
            'can_edit': (application.id_user == request.user) or user_can_edit_all_applications(request.user),
            'application_status': application.status,  # Добавляем статус заявки
            'application_status_display': dict(Application._meta.get_field('status').choices).get(application.status, application.status)
        })
    except Application.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Заявка не найдена'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        
        
@login_required
def check_equipment_availability(request):
    """
    Проверка доступности оборудования на выбранные даты
    - Локационное оборудование: всегда доступно
    - Общее оборудование: учитываем ВСЕ активные заказы (кроме отменённых и завершённых)
    """
    try:
        location_id = request.GET.get('location_id')
        date_start = request.GET.get('date_start')
        date_end = request.GET.get('date_end')
        type_ids = request.GET.getlist('type_ids[]')
        
        print("=" * 60)
        print("ПРОВЕРКА ДОСТУПНОСТИ ОБОРУДОВАНИЯ")
        print(f"Локация ID: {location_id}")
        print(f"Период: {date_start} - {date_end}")
        print(f"Типы: {type_ids}")
        print(f"Текущий пользователь: {request.user.username} (ID: {request.user.id})")
        
        if not location_id:
            return JsonResponse({
                'success': False,
                'error': 'Не указана локация'
            }, status=400)
        
        if not date_start or not date_end:
            return JsonResponse({
                'success': False,
                'error': 'Не указаны даты'
            }, status=400)
        
        # Преобразуем строки в datetime
        start_date = datetime.fromisoformat(date_start)
        end_date = datetime.fromisoformat(date_end)
        
        equipment_list = []
        
        # ========== 1. ЛОКАЦИОННОЕ ОБОРУДОВАНИЕ ==========
        # Всегда показываем всё оборудование как доступное
        equipment_locations = EquipmentLocation.objects.filter(
            id_locations_id=location_id
        )
        
        if type_ids:
            equipment_locations = equipment_locations.filter(
                id_types_equipments_id__in=type_ids
            )
        
        print(f"\nНайдено локационного оборудования: {equipment_locations.count()}")
        
        for eq_loc in equipment_locations:
            total_quantity = eq_loc.quantity
            
            print(f"\nЛокационное оборудование: {eq_loc.id_equipments.name}")
            print(f"  Всего: {total_quantity}")
            print(f"  Доступно: {total_quantity} (всегда доступно)")
            
            equipment_list.append({
                'id': eq_loc.id,
                'equipment_id': eq_loc.id_equipments.id,
                'name': eq_loc.id_equipments.name,
                'type_id': eq_loc.id_types_equipments_id,
                'type_name': eq_loc.id_types_equipments.name,
                'quantity': total_quantity,
                'available': total_quantity,
                'max_quantity': total_quantity,
                'is_common': False
            })
        
        # ========== 2. ОБЩЕЕ ОБОРУДОВАНИЕ ==========
        # Учитываем ВСЕ активные заказы (НЕ отменённые и НЕ завершённые)
        # Активные статусы: 'new', 'in_progress'
        active_statuses = ['new', 'in_progress']
        
        common_equipment = CommonEquipmentLocation.objects.all()
        
        if type_ids:
            common_equipment = common_equipment.filter(
                id_types_equipments_id__in=type_ids
            )
        
        print(f"\nНайдено общего оборудования: {common_equipment.count()}")
        
        for eq in common_equipment:
            total_quantity = eq.quantity
            
            # ========== ПРАВИЛЬНЫЙ РАСЧЁТ ЗАНЯТОСТИ ОБЩЕГО ОБОРУДОВАНИЯ ==========
            # Учитываем ВСЕ заказы (всех пользователей) со статусами 'new' или 'in_progress',
            # которые пересекаются с выбранным периодом
            # Исключаем отменённые и завершённые заявки
            
            # Занятое количество в завершённых заказах (с датами)
            busy_quantity = OrderItem.objects.filter(
                common_equipment_location=eq,
                order__date_time_start__lt=end_date,
                order__date_time_end__gt=start_date,
                order__date_time_end__isnull=False,
                order__id_application__status__in=active_statuses
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            # Активные заказы без даты окончания (бесконечные)
            active_busy = OrderItem.objects.filter(
                common_equipment_location=eq,
                order__date_time_end__isnull=True,
                order__date_time_start__lt=end_date,
                order__id_application__status__in=active_statuses
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            total_busy = busy_quantity + active_busy
            available = total_quantity - total_busy
            
            print(f"\nОбщее оборудование: {eq.id_equipments.name}")
            print(f"  Всего: {total_quantity}")
            print(f"  Занято в активных заказах: {total_busy}")
            print(f"  Доступно: {max(available, 0)}")
            
            equipment_list.append({
                'id': eq.id,
                'equipment_id': eq.id_equipments.id,
                'name': eq.id_equipments.name,
                'type_id': eq.id_types_equipments_id,
                'type_name': eq.id_types_equipments.name,
                'quantity': total_quantity,
                'available': max(available, 0),
                'max_quantity': max(available, 0),
                'is_common': True
            })
        
        print(f"\nИТОГО позиций в ответе: {len(equipment_list)}")
        print("=" * 60)
        
        return JsonResponse({
            'success': True,
            'equipment': equipment_list
        })
        
    except Exception as e:
        print(f"Ошибка в check_equipment_availability: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        
        
@login_required
def room_detail(request, room_id):
    """Страница детального просмотра помещения"""
    location = get_object_or_404(Location, id=room_id)
    
    # Получаем все фото помещения с описанием
    photos = Photo.objects.filter(id_location=location).order_by('-is_main', 'created_at')
    
    # Получаем оборудование в помещении
    equipment_locations = EquipmentLocation.objects.filter(id_locations=location)
    equipment_list = []
    for eq_loc in equipment_locations:
        equipment_list.append({
            'name': eq_loc.id_equipments.name,
            'type': eq_loc.id_types_equipments.name,
            'quantity': eq_loc.quantity
        })
    
    context = {
        'room': location,
        'photos': photos,  # photos содержит поля name, description, photo
        'equipment_list': equipment_list,
        'equipment_count': equipment_locations.count(),
        'total_quantity': sum(eq.quantity for eq in equipment_locations),
    }
    return render(request, 'mainapp/room_detail.html', context)


@login_required
def check_datetime_busy(request):
    """Проверка, занято ли оборудование на выбранные даты"""
    try:
        location_id = request.GET.get('location_id')
        date_start = request.GET.get('date_start')
        date_end = request.GET.get('date_end')
        
        if not location_id or not date_start or not date_end:
            return JsonResponse({
                'busy': False,
                'partially_busy': False,
                'error': 'Не указаны все параметры'
            }, status=400)
        
        start_date = datetime.fromisoformat(date_start)
        end_date = datetime.fromisoformat(date_end)
        
        # Активные статусы: 'new', 'in_progress'
        active_statuses = ['new', 'in_progress']
        
        # Получаем все пересекающиеся заказы (активные, не отменённые, не свои)
        overlapping_orders = Order.objects.filter(
            id_location_id=location_id,
            date_time_start__lt=end_date,
            date_time_end__gt=start_date,
            id_application__status__in=active_statuses
        ).exclude(
            id_user=request.user
        ).select_related('id_application')
        
        # Формируем список занятых интервалов
        busy_intervals = []
        for order in overlapping_orders:
            busy_intervals.append({
                'start': order.date_time_start.isoformat(),
                'end': order.date_time_end.isoformat() if order.date_time_end else None,
                'order_id': order.id,
                'application_name': order.id_application.name if order.id_application else 'Без названия'
            })
        
        # Проверяем оборудование
        equipment_locations = EquipmentLocation.objects.filter(id_locations_id=location_id)
        
        total_available = 0
        total_busy = 0
        
        for eq_loc in equipment_locations:
            busy_quantity = OrderItem.objects.filter(
                equipment_location=eq_loc,
                order__date_time_start__lt=end_date,
                order__date_time_end__gt=start_date,
                order__id_application__status__in=active_statuses
            ).exclude(
                order__id_user=request.user
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            active_busy = OrderItem.objects.filter(
                equipment_location=eq_loc,
                order__date_time_end__isnull=True,
                order__date_time_start__lt=end_date,
                order__id_application__status__in=active_statuses
            ).exclude(
                order__id_user=request.user
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            total_busy += busy_quantity + active_busy
            total_available += eq_loc.quantity
        
        if overlapping_orders.count() > 0:
            return JsonResponse({
                'busy': True,
                'partially_busy': False,
                'busy_count': total_busy,
                'total_available': total_available,
                'busy_orders': busy_intervals,
                'message': 'Помещение занято на выбранные даты'
            })
        elif total_busy > 0 and total_busy >= total_available:
            return JsonResponse({
                'busy': True,
                'partially_busy': False,
                'busy_count': total_busy,
                'total_available': total_available,
                'busy_orders': busy_intervals,
                'message': 'Помещение занято на выбранные даты'
            })
        elif total_busy > 0:
            return JsonResponse({
                'busy': False,
                'partially_busy': True,
                'busy_count': total_busy,
                'total_available': total_available,
                'busy_orders': busy_intervals,
                'message': 'Часть оборудования занята'
            })
        else:
            return JsonResponse({
                'busy': False,
                'partially_busy': False,
                'busy_orders': [],
                'message': 'Помещение доступно'
            })
            
    except Exception as e:
        print(f"Ошибка в check_datetime_busy: {str(e)}")
        return JsonResponse({
            'busy': False,
            'partially_busy': False,
            'error': str(e)
        }, status=400)
        
        
@login_required
def get_busy_time_slots(request):
    """Получить список занятых временных интервалов (исключая отмененные заказы)"""
    try:
        location_id = request.GET.get('location_id')
        
        if not location_id:
            return JsonResponse({
                'success': False,
                'error': 'Не указана локация'
            }, status=400)
        
        busy_slots = []
        
        # Получаем завершенные заказы (исключая отмененные)
        orders = Order.objects.filter(
            id_location_id=location_id,
            date_time_end__isnull=False
        ).exclude(
            id_application__status='cancelled'  # Исключаем отмененные
        )
        
        for order in orders:
            busy_slots.append({
                'start': order.date_time_start.isoformat(),
                'end': order.date_time_end.isoformat(),
                'order_id': order.id
            })
        
        # Активные заказы без даты окончания (исключая отмененные)
        active_orders = Order.objects.filter(
            id_location_id=location_id,
            date_time_end__isnull=True
        ).exclude(
            id_application__status='cancelled'  # Исключаем отмененные
        )
        
        for order in active_orders:
            busy_slots.append({
                'start': order.date_time_start.isoformat(),
                'end': (order.date_time_start + timedelta(days=7)).isoformat(),
                'order_id': order.id,
                'is_active': True
            })
        
        return JsonResponse({
            'success': True,
            'slots': busy_slots
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        

@login_required
def get_busy_dates(request):
    """Получить список занятых дат (исключая отмененные заказы и заказы текущего пользователя)"""
    try:
        location_id = request.GET.get('location_id')
        
        if not location_id:
            return JsonResponse({
                'success': False,
                'error': 'Не указана локация'
            }, status=400)
        
        busy_dates = set()
        
        # Учитываем только НЕ отмененные заказы других пользователей
        orders = Order.objects.filter(
            id_location_id=location_id
        ).exclude(
            date_time_end__isnull=True
        ).exclude(
            id_application__status='cancelled'
        ).exclude(
            id_user=request.user  # Исключаем заказы текущего пользователя
        )
        
        for order in orders:
            if order.date_time_start and order.date_time_end:
                start_date = order.date_time_start.date()
                end_date = order.date_time_end.date()
                
                current = start_date
                while current <= end_date:
                    busy_dates.add(current.isoformat())
                    current = current + timedelta(days=1)
        
        # Активные заказы без даты окончания (исключая отмененные и свои)
        active_orders = Order.objects.filter(
            id_location_id=location_id,
            date_time_end__isnull=True
        ).exclude(
            id_application__status='cancelled'
        ).exclude(
            id_user=request.user  # Исключаем заказы текущего пользователя
        )
        
        for order in active_orders:
            if order.date_time_start:
                start_date = order.date_time_start.date()
                end_date = start_date + timedelta(days=7)
                
                current = start_date
                while current <= end_date:
                    busy_dates.add(current.isoformat())
                    current = current + timedelta(days=1)
        
        print(f"Занятые даты для локации {location_id} (без учета своих заказов): {busy_dates}")
        
        return JsonResponse({
            'success': True,
            'dates': list(busy_dates)
        })
        
    except Exception as e:
        print(f"Ошибка в get_busy_dates: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'dates': []
        }, status=200)


@login_required
def get_applications(request):
    """Получить список заявок текущего пользователя"""
    try:
        applications = Application.objects.filter(id_user=request.user).order_by('-created_at')
        apps_list = []
        
        for app in applications:
            apps_list.append({
                'id': app.id,
                'name': app.name,
                'created_at': app.created_at.isoformat(),
                'date_start': app.date_start.isoformat() if app.date_start else None,  # новое поле
                'date_end': app.date_end.isoformat() if app.date_end else None,        # новое поле
                'status': app.status,
                'comment': app.comment or '',
                'status_display': dict(Application._meta.get_field('status').choices).get(app.status, app.status)
            })
        
        return JsonResponse({
            'success': True,
            'applications': apps_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_application(request):
    """Создать новую заявку"""
    try:
        data = json.loads(request.body)
        name = data.get('name')
        comment = data.get('comment', '')
        
        if not name:
            return JsonResponse({
                'success': False,
                'error': 'Укажите название заявки'
            }, status=400)
        
        application = Application.objects.create(
            name=name,
            id_user=request.user,
            comment=comment,
            status='new'
        )
        
        # Отправляем уведомление о новой заявке
        try:
            EmailService.send_new_application_notification(application, request.user)
        except Exception as e:
            print(f"Ошибка отправки уведомления: {e}")
            # Не блокируем создание заявки, если письмо не отправилось
        
        return JsonResponse({
            'success': True,
            'application': {
                'id': application.id,
                'name': application.name,
                'created_at': application.created_at.isoformat(),
                'status': application.status,
                'comment': application.comment or ''
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def get_application_detail(request, app_id):
    """Получить детали заявки"""
    try:
        application = Application.objects.get(id=app_id, id_user=request.user)
        
        # Получаем связанные заказы
        orders = Order.objects.filter(id_application=application)
        orders_list = []
        
        for order in orders:
            orders_list.append({
                'id': order.id,
                'date_time_start': order.date_time_start.isoformat(),
                'date_time_end': order.date_time_end.isoformat() if order.date_time_end else None,
                'comment': order.comment or '',
                'total_quantity': order.total_quantity
            })
        
        return JsonResponse({
            'success': True,
            'application': {
                'id': application.id,
                'name': application.name,
                'created_at': application.created_at.isoformat(),
                'status': application.status,
                'status_display': dict(Application._meta.get_field('status').choices).get(application.status, application.status),
                'comment': application.comment or '',
                'orders': orders_list
            }
        })
    except Application.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Заявка не найдена'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
        

# ========== РЕДАКТИРОВАНИЕ ЗАКАЗОВ ==========

@login_required
def get_order_details(request, order_id):
    """Получение деталей заказа для редактирования с проверкой прав (группировка слотов)"""
    try:
        order = Order.objects.get(id=order_id)
        
        # Проверка прав доступа к заказу
        if order.id_user != request.user and not user_can_edit_all_applications(request.user):
            return JsonResponse({
                'success': False, 
                'error': 'У вас нет прав для редактирования этого заказа'
            }, status=403)
        
        items = OrderItem.objects.filter(order=order)
        
        # Разделяем на обычные позиции и слоты
        regular_equipment = []
        slots_map = {}  # Используем словарь для группировки слотов по времени
        
        for item in items:
            if item.is_slot:
                # Ключ для группировки - дата и время слота
                if item.slot_date_start and item.slot_date_end:
                    slot_key = f"{item.slot_date_start.isoformat()}|{item.slot_date_end.isoformat()}"
                    
                    if slot_key not in slots_map:
                        slots_map[slot_key] = {
                            'date_start': item.slot_date_start.isoformat(),
                            'date_end': item.slot_date_end.isoformat(),
                            'equipment': []
                        }
                    
                    # Добавляем оборудование в слот
                    if item.equipment_location:
                        slots_map[slot_key]['equipment'].append({
                            'equipment_id': item.equipment_location.id_equipments.id,
                            'equipment_name': item.equipment_location.id_equipments.name,
                            'type_name': item.equipment_location.id_types_equipments.name,
                            'quantity': item.quantity,
                            'max_quantity': item.equipment_location.quantity,
                            'is_common': False,
                            'order_item_id': item.id
                        })
                    elif item.common_equipment_location:
                        slots_map[slot_key]['equipment'].append({
                            'equipment_id': item.common_equipment_location.id_equipments.id,
                            'equipment_name': item.common_equipment_location.id_equipments.name,
                            'type_name': item.common_equipment_location.id_types_equipments.name,
                            'quantity': item.quantity,
                            'max_quantity': item.common_equipment_location.quantity,
                            'is_common': True,
                            'order_item_id': item.id
                        })
            else:
                # Обычная позиция (не слот)
                if item.equipment_location:
                    regular_equipment.append({
                        'equipment_id': item.equipment_location.id_equipments.id,
                        'equipment_name': item.equipment_location.id_equipments.name,
                        'type_name': item.equipment_location.id_types_equipments.name,
                        'quantity': item.quantity,
                        'is_common': False,
                        'max_quantity': item.equipment_location.quantity,
                        'order_item_id': item.id
                    })
                elif item.common_equipment_location:
                    regular_equipment.append({
                        'equipment_id': item.common_equipment_location.id_equipments.id,
                        'equipment_name': item.common_equipment_location.id_equipments.name,
                        'type_name': item.common_equipment_location.id_types_equipments.name,
                        'quantity': item.quantity,
                        'is_common': True,
                        'max_quantity': item.common_equipment_location.quantity,
                        'order_item_id': item.id
                    })
        
        # Преобразуем словарь в список и сортируем по дате начала
        slots_list = list(slots_map.values())
        slots_list.sort(key=lambda x: x['date_start'])
        
        return JsonResponse({
            'success': True,
            'order': {
                'id': order.id,
                'location_id': order.id_location.id if order.id_location else None,
                'location_name': order.id_location.name if order.id_location else 'Общее оборудование',
                'date_time_start': order.date_time_start.isoformat() if order.date_time_start else None,
                'date_time_end': order.date_time_end.isoformat() if order.date_time_end else None,
                'comment': order.comment or '',
                'equipment': regular_equipment,
                'slots': slots_list,  # Теперь это сгруппированные слоты
                'has_slots': len(slots_list) > 0,
                'can_edit': (order.id_user == request.user) or user_can_edit_all_applications(request.user)
            }
        })
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Заказ не найден'}, status=404)
    except Exception as e:
        print(f"Ошибка в get_order_details: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_order(request, order_id):
    """Обновление заказа с проверкой прав (поддержка слотов)"""
    try:
        data = json.loads(request.body)
        order = Order.objects.get(id=order_id)
        
        # Проверка прав на редактирование
        if order.id_user != request.user and not user_can_edit_all_applications(request.user):
            return JsonResponse({
                'success': False,
                'error': 'У вас нет прав для редактирования этого заказа'
            }, status=403)
        
        if order.id_application and order.id_application.status in ['completed', 'cancelled']:
            return JsonResponse({
                'success': False,
                'error': 'Нельзя редактировать завершенную или отмененную заявку'
            }, status=400)
        
        with transaction.atomic():
            # Обновляем основные поля заказа
            if 'date_time_start' in data:
                order.date_time_start = data['date_time_start']
            if 'date_time_end' in data:
                order.date_time_end = data['date_time_end']
            if 'comment' in data:
                order.comment = data['comment']
            order.save()
            
            # Удаляем старые позиции
            OrderItem.objects.filter(order=order).delete()
            
            items_added = 0
            
            # ========== ОБРАБОТКА СЛОТОВ ==========
            slots_data = data.get('slots', [])
            is_slots_order = data.get('type') == 'slots' or len(slots_data) > 0
            
            if is_slots_order and slots_data:
                for slot in slots_data:
                    slot_date_start = slot.get('date_start')
                    slot_date_end = slot.get('date_end')
                    slot_equipment = slot.get('equipment', [])
                    
                    if not slot_date_start or not slot_date_end:
                        continue
                    
                    for eq in slot_equipment:
                        equipment_id = eq.get('equipment_id')
                        quantity = eq.get('quantity', 0)
                        is_common = eq.get('is_common', False)
                        
                        if quantity <= 0:
                            continue
                        
                        if is_common:
                            common_equipment = CommonEquipmentLocation.objects.filter(
                                id_equipments_id=equipment_id
                            ).first()
                            if common_equipment:
                                OrderItem.objects.create(
                                    order=order,
                                    common_equipment_location=common_equipment,
                                    quantity=min(quantity, common_equipment.quantity),
                                    is_slot=True,
                                    slot_date_start=slot_date_start,
                                    slot_date_end=slot_date_end
                                )
                                items_added += 1
                        else:
                            if order.id_location:
                                equipment_location = EquipmentLocation.objects.filter(
                                    id_locations_id=order.id_location.id,
                                    id_equipments_id=equipment_id
                                ).first()
                                if equipment_location:
                                    OrderItem.objects.create(
                                        order=order,
                                        equipment_location=equipment_location,
                                        quantity=min(quantity, equipment_location.quantity),
                                        is_slot=True,
                                        slot_date_start=slot_date_start,
                                        slot_date_end=slot_date_end
                                    )
                                    items_added += 1
            
            # ========== ОБЫЧНОЕ ОБОРУДОВАНИЕ (НЕ СЛОТЫ) ==========
            equipment_data = data.get('equipment', [])
            for eq_data in equipment_data:
                quantity = eq_data.get('quantity', 0)
                if quantity <= 0:
                    continue
                
                if eq_data.get('is_common', False):
                    common_equipment = CommonEquipmentLocation.objects.filter(
                        id_equipments_id=eq_data['equipment_id']
                    ).first()
                    if common_equipment:
                        OrderItem.objects.create(
                            order=order,
                            common_equipment_location=common_equipment,
                            quantity=min(quantity, common_equipment.quantity),
                            is_slot=False
                        )
                        items_added += 1
                else:
                    if order.id_location:
                        equipment_location = EquipmentLocation.objects.filter(
                            id_locations_id=order.id_location.id,
                            id_equipments_id=eq_data['equipment_id']
                        ).first()
                        if equipment_location:
                            OrderItem.objects.create(
                                order=order,
                                equipment_location=equipment_location,
                                quantity=min(quantity, equipment_location.quantity),
                                is_slot=False
                            )
                            items_added += 1
            
            # Если нет ни одной позиции, удаляем заказ
            if items_added == 0:
                order.delete()
                return JsonResponse({
                    'success': False,
                    'error': 'Не добавлено ни одной позиции'
                }, status=400)
            
            # Обновляем согласования для заявки
            if order.id_application:
                # Собираем типы оборудования из обновленного заказа
                equipment_types = set()
                new_items = OrderItem.objects.filter(order=order)
                for item in new_items:
                    if item.equipment_location:
                        equipment_types.add(item.equipment_location.id_types_equipments.id)
                    elif item.common_equipment_location:
                        equipment_types.add(item.common_equipment_location.id_types_equipments.id)
                
                # Обновляем согласования
                from .views_helpers import update_application_approvals_for_types
                update_application_approvals_for_types(order.id_application, equipment_types)
            
            HistoryService.add_entry(
                request.user,
                f"Редактирован заказ #{order.id} для локации {order.id_location.name if order.id_location else 'Общее оборудование'} (слотов: {len(slots_data)})"
            )
        
        return JsonResponse({'success': True, 'message': 'Заказ успешно обновлен', 'items_added': items_added})
    
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Заказ не найден'}, status=404)
    except Exception as e:
        print(f"Ошибка в update_order: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def cancel_order(request, order_id):
    """Отмена заказа с проверкой прав"""
    try:
        order = Order.objects.get(id=order_id)
        
        if order.id_user != request.user and not user_can_edit_all_applications(request.user):
            return JsonResponse({
                'success': False,
                'error': 'У вас нет прав для отмены этого заказа'
            }, status=403)
        
        if order.id_application:
            if order.id_application.status == 'completed':
                return JsonResponse({
                    'success': False,
                    'error': 'Нельзя отменить завершенную заявку'
                }, status=400)
            
            order.id_application.status = 'cancelled'
            order.id_application.save()
            
            HistoryService.add_entry(
                request.user,
                f"Отменена заявка #{order.id_application.id} - {order.id_application.name[:50]}"
            )
        
        return JsonResponse({'success': True, 'message': 'Заявка отменена'})
    
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Заказ не найден'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def duplicate_order(request, order_id):
    """Дублирование заказа с созданием согласований для соответствующих подразделений"""
    try:
        original_order = Order.objects.get(id=order_id, id_user=request.user)
        
        with transaction.atomic():
            new_application = Application.objects.create(
                name=f"Копия: {original_order.id_application.name if original_order.id_application else 'Заказ'}",
                id_user=request.user,
                comment=f"Создано из заказа #{original_order.id}",
                status='new'
            )
            
            new_order = Order.objects.create(
                id_user=request.user,
                id_application=new_application,
                id_location=original_order.id_location,
                date_time_start=original_order.date_time_start,
                date_time_end=original_order.date_time_end,
                comment=original_order.comment
            )
            
            # Копируем оборудование и собираем типы
            equipment_types_in_order = set()
            
            for item in OrderItem.objects.filter(order=original_order):
                if item.equipment_location:
                    OrderItem.objects.create(
                        order=new_order,
                        equipment_location=item.equipment_location,
                        quantity=item.quantity
                    )
                    equipment_types_in_order.add(item.equipment_location.id_types_equipments.id)
                elif item.common_equipment_location:
                    OrderItem.objects.create(
                        order=new_order,
                        common_equipment_location=item.common_equipment_location,
                        quantity=item.quantity
                    )
                    equipment_types_in_order.add(item.common_equipment_location.id_types_equipments.id)
            
            # Создаем согласования для подразделений, связанных с типами оборудования
            departments_for_approval = Department.objects.filter(
                is_approval_required=True
            ).filter(
                department_type_equipment__id_type_equipment__id__in=equipment_types_in_order
            ).distinct()
            
            for dept in departments_for_approval:
                ApplicationApproval.objects.get_or_create(
                    id_department=dept,
                    id_application=new_application,
                    defaults={
                        'is_agreed': False,
                        'comment': f'Ожидает согласования от подразделения {dept.name}'
                    }
                )
            
            HistoryService.add_entry(
                request.user,
                f"Создана копия заказа #{original_order.id} -> #{new_order.id}"
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Копия заказа создана',
            'new_order_id': new_order.id
        })
    
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Заказ не найден'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_order_item_approval(request, item_id):
    """
    Обновление данных согласования для позиции заказа
    """
    try:
        data = json.loads(request.body)
        can_provide = data.get('can_provide', 0)
        is_agreed = data.get('is_agreed', False)
        
        # Получаем позицию заказа
        order_item = OrderItem.objects.get(id=item_id)
        
        # Проверяем права: пользователь должен быть в подразделении,
        # которое связано с типом оборудования этой позиции
        order = order_item.order
        location = order.id_location
        
        # Получаем тип оборудования позиции
        if order_item.equipment_location:
            equipment_type = order_item.equipment_location.id_types_equipments
        elif order_item.common_equipment_location:
            equipment_type = order_item.common_equipment_location.id_types_equipments
        else:
            return JsonResponse({
                'success': False,
                'error': 'Не удалось определить тип оборудования'
            }, status=400)
        
        # Проверяем, принадлежит ли пользователь к подразделению,
        # которое может согласовывать этот тип оборудования
        user_departments = UserDepartment.objects.filter(id_user=request.user)
        department_ids = [ud.id_department.id for ud in user_departments]
        
        # Проверяем, есть ли связь между подразделениями пользователя и типом оборудования
        can_approve = DepartmentTypeEquipment.objects.filter(
            id_department_id__in=department_ids,
            id_type_equipment=equipment_type
        ).exists()
        
        if not can_approve:
            return JsonResponse({
                'success': False,
                'error': 'У вас нет прав для согласования этого оборудования'
            }, status=403)
        
        # Обновляем поля
        order_item.can_provide = can_provide
        order_item.is_agreed = is_agreed
        order_item.save()
        
        # Логируем действие
        HistoryService.add_entry(
            request.user,
            f"Обновлено согласование для позиции #{order_item.id}: количество={can_provide}, согласовано={is_agreed}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Данные согласования сохранены',
            'order_item': {
                'id': order_item.id,
                'can_provide': order_item.can_provide,
                'is_agreed': order_item.is_agreed
            }
        })
        
    except OrderItem.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Позиция заказа не найдена'
        }, status=404)
    except Exception as e:
        print(f"Ошибка обновления согласования: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
    
# ========== ПОДРАЗДЕЛЕНИЯ (DEPARTMENT) - ТОЛЬКО ПРОСМОТР ==========

@login_required
def get_departments(request):
    """
    Получить список всех подразделений
    """
    try:
        departments = Department.objects.all()
        
        limit = int(request.GET.get('limit', 100))
        offset = int(request.GET.get('offset', 0))
        
        total = departments.count()
        departments_page = departments[offset:offset+limit]
        
        return JsonResponse({
            'success': True,
            'departments': [
                {
                    'id': d.id,
                    'name': d.name,
                    'description': d.description,
                    'user_count': d.department_users.count(),
                    'created_at': d.created_at.isoformat(),
                    'created_at_formatted': d.created_at.strftime('%d.%m.%Y %H:%M:%S'),
                    'updated_at': d.updated_at.isoformat(),
                    'updated_at_formatted': d.updated_at.strftime('%d.%m.%Y %H:%M:%S')
                }
                for d in departments_page
            ],
            'total': total,
            'limit': limit,
            'offset': offset
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_department_detail(request, dept_id):
    """
    Получить детали подразделения по ID
    """
    try:
        department = Department.objects.get(id=dept_id)
        
        # Получаем пользователей подразделения
        users = []
        for ud in department.department_users.all():
            users.append({
                'user_id': ud.id_user.id,
                'username': ud.id_user.username,
                'email': ud.id_user.email,
                'full_name': ud.id_user.get_full_name() or ud.id_user.username,
                'is_head': ud.is_head
            })
        
        return JsonResponse({
            'success': True,
            'department': {
                'id': department.id,
                'name': department.name,
                'description': department.description,
                'user_count': len(users),
                'users': users,
                'created_at': department.created_at.isoformat(),
                'created_at_formatted': department.created_at.strftime('%d.%m.%Y %H:%M:%S'),
                'updated_at': department.updated_at.isoformat(),
                'updated_at_formatted': department.updated_at.strftime('%d.%m.%Y %H:%M:%S')
            }
        })
        
    except Department.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Подразделение не найдено'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_user_departments(request):
    """
    Получить подразделения текущего пользователя
    """
    try:
        user_departments = UserDepartment.objects.filter(id_user=request.user)
        
        departments = []
        for ud in user_departments:
            departments.append({
                'id': ud.id_department.id,
                'name': ud.id_department.name,
                'description': ud.id_department.description,
                'is_head': ud.is_head,
                'joined_at': ud.created_at.isoformat(),
                'joined_at_formatted': ud.created_at.strftime('%d.%m.%Y %H:%M:%S')
            })
        
        return JsonResponse({
            'success': True,
            'departments': departments,
            'total': len(departments)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        

@login_required
def get_user_departments_by_id(request, user_id):
    """
    Получить подразделения пользователя по ID
    (только для администраторов или для своего пользователя)
    """
    try:
        # Проверка прав: можно смотреть только свои подразделения или админу
        if request.user.id != user_id and not request.user.is_superuser and not user_can_view_all_applications(request.user):
            return JsonResponse({
                'success': False,
                'error': 'У вас нет прав для просмотра подразделений другого пользователя'
            }, status=403)
        
        user = User.objects.get(id=user_id)
        user_departments = UserDepartment.objects.filter(id_user=user)
        
        departments = []
        for ud in user_departments:
            departments.append({
                'id': ud.id_department.id,
                'name': ud.id_department.name,
                'description': ud.id_department.description,
                'is_head': ud.is_head,
                'joined_at': ud.created_at.isoformat(),
                'joined_at_formatted': ud.created_at.strftime('%d.%m.%Y %H:%M:%S')
            })
        
        return JsonResponse({
            'success': True,
            'user_id': user.id,
            'username': user.username,
            'departments': departments,
            'total': len(departments)
        })
        
    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Пользователь не найден'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
                
@login_required
def get_user_department(request):
    """
    Получить подразделение текущего пользователя
    """
    try:
        department = Department.objects.filter(id_user=request.user).first()
        
        if not department:
            return JsonResponse({
                'success': True,
                'department': None,
                'message': 'Пользователь не привязан к подразделению'
            })
        
        return JsonResponse({
            'success': True,
            'department': {
                'id': department.id,
                'name': department.name,
                'created_at': department.created_at.isoformat(),
                'created_at_formatted': department.created_at.strftime('%d.%m.%Y %H:%M:%S')
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    

def create_application_approvals(application, equipment_types):
    """
    Создает записи согласования для заявки на основе типов оборудования
    """
    if not equipment_types:
        print(f"Нет типов оборудования для заявки #{application.id}")
        return 0
    
    departments = get_departments_for_equipment_types(equipment_types)
    
    print(f"Подразделения для согласования заявки #{application.id}: {[d.name for d in departments]}")
    
    created_count = 0
    for dept in departments:
        approval, created = ApplicationApproval.objects.get_or_create(
            id_department=dept,
            id_application=application,
            defaults={
                'is_agreed': False,
                'comment': f'Ожидает согласования от подразделения {dept.name}'
            }
        )
        if created:
            created_count += 1
            print(f"Создана запись согласования для подразделения '{dept.name}'")
    
    return created_count


@login_required
def get_user_department_types(request):
    """
    Получить типы оборудования, связанные с подразделениями пользователя
    """
    try:
        # Получаем все подразделения пользователя
        user_departments = UserDepartment.objects.filter(id_user=request.user)
        
        if not user_departments.exists():
            return JsonResponse({
                'success': True,
                'department_types': [],
                'message': 'Пользователь не привязан ни к одному подразделению'
            })
        
        # Получаем ID подразделений пользователя
        department_ids = [ud.id_department.id for ud in user_departments]
        
        # Получаем все связи подразделений с типами оборудования
        department_type_relations = DepartmentTypeEquipment.objects.filter(
            id_department_id__in=department_ids
        ).select_related('id_type_equipment')
        
        # Формируем список уникальных типов оборудования
        department_types = []
        seen_types = set()
        
        for relation in department_type_relations:
            type_name = relation.id_type_equipment.name
            if type_name not in seen_types:
                seen_types.add(type_name)
                department_types.append({
                    'id': relation.id_type_equipment.id,
                    'name': type_name,
                    'department_id': relation.id_department.id,
                    'department_name': relation.id_department.name
                })
        
        return JsonResponse({
            'success': True,
            'department_types': department_types,
            'total': len(department_types)
        })
        
    except Exception as e:
        print(f"Ошибка в get_user_department_types: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def approve_application(request):
    """
    Согласование заявки от имени подразделения
    """
    try:
        data = json.loads(request.body)
        application_id = data.get('application_id')
        department_id = data.get('department_id')
        
        if not application_id or not department_id:
            return JsonResponse({
                'success': False,
                'error': 'Не указаны параметры'
            }, status=400)
        
        # Получаем заявку
        application = Application.objects.get(id=application_id)
        
        # Получаем подразделение
        department = Department.objects.get(id=department_id)
        
        # Проверяем, принадлежит ли пользователь к этому подразделению
        user_in_department = UserDepartment.objects.filter(
            id_user=request.user,
            id_department=department
        ).exists()
        
        if not user_in_department and not request.user.is_superuser:
            return JsonResponse({
                'success': False,
                'error': 'Вы не принадлежите к этому подразделению'
            }, status=403)
        
        # Получаем или создаем запись согласования
        approval, created = ApplicationApproval.objects.get_or_create(
            id_department=department,
            id_application=application,
            defaults={
                'is_agreed': False,
                'comment': ''
            }
        )
        
        # Если уже согласовано
        if approval.is_agreed:
            return JsonResponse({
                'success': False,
                'error': 'Эта заявка уже согласована вашим подразделением'
            }, status=400)
        
        # Обновляем статус согласования
        approval.is_agreed = True
        approval.id_user = request.user
        approval.date_agreed = timezone.now()
        approval.comment = f"Согласовано пользователем {request.user.get_full_name() or request.user.username}"
        approval.save()
        
        # ========== ПРОВЕРЯЕМ, ВСЕ ЛИ ПОДРАЗДЕЛЕНИЯ СОГЛАСОВАЛИ ==========
        # Получаем все подразделения, которые должны согласовать эту заявку
        # (подразделения с is_approval_required=True, связанные с типами оборудования)
        
        # Получаем все заказы в заявке
        orders = Order.objects.filter(id_application=application)
        
        # Собираем все типы оборудования из заявки
        equipment_types = set()
        for order in orders:
            items = OrderItem.objects.filter(order=order)
            for item in items:
                if item.equipment_location:
                    equipment_types.add(item.equipment_location.id_types_equipments.id)
                elif item.common_equipment_location:
                    equipment_types.add(item.common_equipment_location.id_types_equipments.id)
        
        # Находим все подразделения, которые должны согласовать
        required_departments = Department.objects.filter(
            is_approval_required=True
        ).filter(
            department_type_equipment__id_type_equipment__id__in=equipment_types
        ).distinct()
        
        # Получаем все согласования для этой заявки
        approvals = ApplicationApproval.objects.filter(id_application=application)
        
        # Проверяем, все ли обязательные подразделения согласовали
        required_dept_ids = set(required_departments.values_list('id', flat=True))
        approved_dept_ids = set(approvals.filter(is_agreed=True).values_list('id_department_id', flat=True))
        
        # Если все обязательные подразделения согласовали, меняем статус заявки
        if required_dept_ids.issubset(approved_dept_ids):
            application.status = 'in_progress'
            application.save(update_fields=['status'])
            
            # Логируем смену статуса
            HistoryService.add_entry(
                request.user,
                f"Статус заявки #{application.id} изменён на 'В работе' (все подразделения согласовали)"
            )
            
            # Отправляем уведомление о смене статуса
            try:
                from .email_utils import EmailService
                EmailService.send_application_status_notification(
                    application, 
                    'new', 
                    'in_progress', 
                    request.user.username
                )
            except Exception as e:
                print(f"Ошибка отправки уведомления: {e}")
        
        # Логируем действие
        HistoryService.add_entry(
            request.user,
            f"Согласована заявка #{application.id} от подразделения {department.name}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Заявка #{application.id} успешно согласована',
            'approval': {
                'id': approval.id,
                'is_agreed': approval.is_agreed,
                'date_agreed': approval.date_agreed.isoformat(),
                'user_name': request.user.username
            },
            'application_status': application.status,
            'application_status_display': dict(Application._meta.get_field('status').choices).get(application.status, application.status),
            'all_approved': required_dept_ids.issubset(approved_dept_ids)
        })
        
    except Application.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Заявка не найдена'
        }, status=404)
    except Department.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Подразделение не найдено'
        }, status=404)
    except Exception as e:
        print(f"Ошибка согласования заявки: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)